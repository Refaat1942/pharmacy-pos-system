"""Shared Excel (.xlsx) export helpers."""
from io import BytesIO

from fastapi.responses import StreamingResponse


def xlsx_safe(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return v
    return str(v)


def xlsx_response(headers: list, rows: list, filename: str) -> StreamingResponse:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append([xlsx_safe(h) for h in headers])
    for row in rows:
        ws.append([xlsx_safe(c) for c in row])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def xlsx_multi_sheet(
    sheets: list[tuple[str, list, list]],
    filename: str,
) -> StreamingResponse:
    """sheets: [(sheet_name, headers, rows), ...]"""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for title, headers, rows in sheets:
        ws = wb.create_sheet(title=title[:31])
        ws.append([xlsx_safe(h) for h in headers])
        for row in rows:
            ws.append([xlsx_safe(c) for c in row])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
