"""Customer accounts module — CRUD, payments, statements with running balance."""
import csv
import io
import re
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
import psycopg2.extras
from db import get_db_connection
from deps import get_current_user
from regions import resolve_region_key, REGIONS

router = APIRouter(prefix="/api", tags=["customers"])

# Back-compat alias — prefer digital_platforms module.
from digital_platforms import (
    is_platform_partner_customer_id,
    lookup_platform_partner,
    platform_display_name as platform_partner_display_name,
    sql_exclude_platform_partner_customers,
)

PLATFORM_PARTNER_NAMES = {
    "talabat": "Talabat",
    "vezeeta": "Vezeeta",
    "other_digital": "Other Digital",
}


def _admin_only(user):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admins only")


def _row_get(r, *keys):
    for k in keys:
        if k in r and r[k] not in (None, ""):
            return r[k]
    return None


def _norm_phone(phone: str | None) -> str | None:
    if not phone:
        return None
    digits = re.sub(r"\D", "", str(phone).strip())
    return digits or None


def _parse_bool(v) -> bool:
    if v in (None, ""):
        return True
    s = str(v).strip().lower()
    if s in ("false", "no", "0", "n", "inactive"):
        return False
    return True


def _norm_customer_code(code: str | None) -> str | None:
    if code in (None, ""):
        return None
    c = str(code).strip().upper()
    return c or None


def _default_customer_code(customer_id: int) -> str:
    return f"C{customer_id:06d}"


def _ensure_customer_code(cur, customer_id: int, code: str | None) -> str:
    norm = _norm_customer_code(code)
    if norm:
        cur.execute(
            "SELECT id FROM customers WHERE UPPER(TRIM(code)) = %s AND id <> %s LIMIT 1",
            (norm, customer_id),
        )
        if cur.fetchone():
            raise HTTPException(status_code=400, detail=f"Customer code '{norm}' already in use")
        return norm
    return _default_customer_code(customer_id)


def _parse_branch_ids(raw, branch_by_id: dict[int, dict], branch_by_name: dict[str, int]) -> list[int]:
    if raw in (None, ""):
        return []
    ids: list[int] = []
    for part in str(raw).split(","):
        token = part.strip()
        if not token:
            continue
        if token.isdigit():
            bid = int(token)
            if bid in branch_by_id:
                ids.append(bid)
            continue
        key = token.lower()
        if key in branch_by_name:
            ids.append(branch_by_name[key])
    return list(dict.fromkeys(ids))


CUSTOMER_TEMPLATE_HEADERS = [
    "Code", "Name", "Phone", "Email", "Region", "Address Details",
    "Tax #", "Credit Limit", "Notes", "Active", "Authorized Branches",
]


@router.get("/customers/bulk-template")
def customers_bulk_template(current_user=Depends(get_current_user)):
    """Download a blank Excel template for bulk customer upload."""
    _admin_only(current_user)
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append(CUSTOMER_TEMPLATE_HEADERS)
    ws.append([
        "C000001",
        "Ahmed Pharmacy Co.",
        "01001234567",
        "ahmed@example.com",
        "Ismailia City",
        "12 El Gomhoria St, Building 3, Apt 5",
        "123-456-789",
        5000,
        "Wholesale account",
        "yes",
        "Main Branch, Branch 2",
    ])
    ws.append([
        "Example Customer (minimal)",
        "",
        "",
        "fayed",
        "",
        "",
        0,
        "",
        "true",
        "",
    ])

    regions_ws = wb.create_sheet("Regions (reference)")
    regions_ws.append(["Region Key", "English", "Arabic"])
    for r in REGIONS:
        regions_ws.append([r["key"], r["en"], r["ar"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="customers_template.xlsx"'},
    )


@router.post("/customers/bulk-upload")
async def customers_bulk_upload(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
):
    """Bulk import customers from Excel/CSV. Matches by ID or phone to update."""
    _admin_only(current_user)
    from openpyxl import load_workbook

    content = await file.read()
    name = (file.filename or "").lower()

    rows: list[dict] = []
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    else:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(r):
                continue
            row = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
            rows.append(row)

    rows = [
        {(str(k).strip().lower() if k is not None else ""): v for k, v in row.items()}
        for row in rows
    ]

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id, name_en, name_ar FROM branches")
    branches = [dict(b) for b in cur.fetchall()]
    branch_by_id = {b["id"]: b for b in branches}
    branch_by_name: dict[str, int] = {}
    for b in branches:
        for field in ("name_en", "name_ar"):
            val = (b.get(field) or "").strip().lower()
            if val:
                branch_by_name[val] = b["id"]

    inserted = updated = errors = 0
    error_details: list[str] = []
    user_id = current_user.get("user_id")

    for idx, r in enumerate(rows, start=2):
        cur.execute("SAVEPOINT row_sp")
        try:
            cust_id_raw = _row_get(r, "id", "customer id", "customer_id")
            name_val = str(_row_get(r, "name", "customer name", "customer") or "").strip()
            if not name_val:
                raise ValueError("Name is required")

            phone_raw = _row_get(r, "phone", "mobile", "tel", "telephone")
            phone = str(phone_raw).strip() if phone_raw not in (None, "") else None
            email_raw = _row_get(r, "email", "e-mail")
            email = str(email_raw).strip() if email_raw not in (None, "") else None
            region_raw = _row_get(r, "region", "area", "area / region")
            region = resolve_region_key(customer_region=str(region_raw).strip() if region_raw else None)
            if region == "unknown" and region_raw not in (None, ""):
                region = str(region_raw).strip()
            elif region == "unknown":
                region = None

            address = str(_row_get(r, "address details", "address_details", "address") or "").strip() or None
            tax_number = str(_row_get(r, "tax #", "tax", "tax_number", "tax number") or "").strip() or None
            credit_limit = float(_row_get(r, "credit limit", "credit_limit", "limit") or 0)
            notes = str(_row_get(r, "notes", "note") or "").strip() or None
            active = _parse_bool(_row_get(r, "active", "status"))
            branch_ids = _parse_branch_ids(
                _row_get(r, "authorized branches", "authorized_branches", "branches", "branch ids"),
                branch_by_id,
                branch_by_name,
            )

            code_raw = _row_get(r, "code", "customer code", "customer_code")
            code_val = _norm_customer_code(str(code_raw).strip() if code_raw not in (None, "") else None)

            existing = None
            if cust_id_raw not in (None, ""):
                try:
                    cid = int(float(str(cust_id_raw).strip()))
                except ValueError as exc:
                    raise ValueError("Invalid customer ID") from exc
                cur.execute("SELECT id FROM customers WHERE id=%s", (cid,))
                existing = cur.fetchone()
                if not existing:
                    raise ValueError(f"Customer ID {cid} not found")

            if not existing and code_val:
                cur.execute(
                    "SELECT id FROM customers WHERE UPPER(TRIM(code)) = %s ORDER BY id LIMIT 1",
                    (code_val,),
                )
                existing = cur.fetchone()

            if not existing and phone:
                norm = _norm_phone(phone)
                if norm:
                    cur.execute(
                        "SELECT id FROM customers WHERE "
                        "REGEXP_REPLACE(COALESCE(phone, ''), '[^0-9]', '', 'g') = %s "
                        "ORDER BY id LIMIT 1",
                        (norm,),
                    )
                    existing = cur.fetchone()

            if existing:
                cid = existing["id"]
                final_code = _ensure_customer_code(cur, cid, code_val)
                cur.execute(
                    """UPDATE customers SET code=%s, name=%s, phone=%s, email=%s, region=%s,
                       address_details=%s, tax_number=%s, credit_limit=%s, notes=%s, active=%s
                       WHERE id=%s""",
                    (final_code, name_val, phone, email, region, address, tax_number,
                     credit_limit, notes, active, cid),
                )
                if branch_ids:
                    cur.execute("DELETE FROM customer_branches WHERE customer_id=%s", (cid,))
                    for bid in branch_ids:
                        cur.execute(
                            "INSERT INTO customer_branches (customer_id, branch_id, authorized_by) "
                            "VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                            (cid, bid, user_id),
                        )
                updated += 1
            else:
                cur.execute(
                    """INSERT INTO customers
                       (name, phone, email, region, address_details, tax_number,
                        credit_limit, notes, active)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                    (name_val, phone, email, region, address, tax_number,
                     credit_limit, notes, active),
                )
                cid = cur.fetchone()["id"]
                final_code = _ensure_customer_code(cur, cid, code_val)
                cur.execute("UPDATE customers SET code=%s WHERE id=%s", (final_code, cid))
                for bid in branch_ids:
                    cur.execute(
                        "INSERT INTO customer_branches (customer_id, branch_id, authorized_by) "
                        "VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                        (cid, bid, user_id),
                    )
                inserted += 1

            cur.execute("RELEASE SAVEPOINT row_sp")
        except Exception as e:
            cur.execute("ROLLBACK TO SAVEPOINT row_sp")
            errors += 1
            error_details.append(f"Row {idx}: {e}")

    conn.commit()
    conn.close()
    return {
        "inserted": inserted,
        "updated": updated,
        "errors": errors,
        "error_details": error_details[:50],
    }


class CustomerIn(BaseModel):
    name: str
    code: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    region: Optional[str] = None
    address_details: Optional[str] = None
    tax_number: Optional[str] = None
    credit_limit: float = 0
    notes: Optional[str] = None
    discount_percent: Optional[float] = None
    discount_notes: Optional[str] = None
    active: bool = True
    branch_ids: Optional[list[int]] = None
    sale_type: Optional[str] = "cash"


@router.get("/customers/v2")
def list_customers_v2(
    q: str = "",
    active_only: bool = True,
    include_platform_partners: bool = False,
    current_user=Depends(get_current_user),
):
    """Extended customer list with balance, charged, paid, region.

  Digital platform B2B partners (Talabat, Instashop, etc.) are excluded by default;
  manage them under Settings → Digital Platforms.
    """
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    where = []
    params: list = []
    if not include_platform_partners:
        where.append(sql_exclude_platform_partner_customers("c"))
    if active_only:
        where.append("c.active = true")
    if q:
        where.append("(c.name ILIKE %s OR c.phone ILIKE %s OR c.email ILIKE %s OR c.code ILIKE %s)")
        like = f"%{q}%"
        params += [like, like, like, like]
    # Non-admin: row-level branch-scope — only customers with at least one
    # invoice in the user's branch are visible; aggregates also branch-scoped.
    if current_user.get("role") != "admin":
        ub = current_user.get("branch_id")
        if ub is None:
            raise HTTPException(status_code=403, detail="No branch assigned")
        where.append("EXISTS (SELECT 1 FROM customer_branches cb WHERE cb.customer_id=c.id AND cb.branch_id=%s)")
        charged_sub = ("(SELECT COALESCE(SUM(net_total),0) FROM invoices "
                       "WHERE customer_id=c.id AND payment_method='account' "
                       "AND type!='return' AND branch_id=%s)")
        paid_sub = ("(SELECT COALESCE(SUM(cp.amount),0) FROM customer_payments cp "
                    "JOIN invoices i ON cp.invoice_id=i.id "
                    "WHERE cp.customer_id=c.id AND i.branch_id=%s)")
        params = [ub, ub] + params + [ub]
    else:
        charged_sub = ("(SELECT COALESCE(SUM(net_total),0) FROM invoices "
                       "WHERE customer_id=c.id AND payment_method='account' AND type!='return')")
        paid_sub = ("(SELECT COALESCE(SUM(amount),0) FROM customer_payments "
                    "WHERE customer_id=c.id)")
    sql = f"SELECT c.*, {charged_sub} AS total_charged, {paid_sub} AS total_paid FROM customers c"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY c.name ASC LIMIT 500"
    cur.execute(sql, params)
    rows = [dict(r) for r in cur.fetchall()]
    for r in rows:
        r["balance"] = float(r["total_charged"]) - float(r["total_paid"])
    conn.close()
    return rows


@router.post("/customers/v2")
def create_customer_v2(req: CustomerIn, current_user=Depends(get_current_user)):
    _admin_only(current_user)
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            """INSERT INTO customers
               (name, phone, email, region, address_details, tax_number,
                credit_limit, notes, discount_percent, discount_notes, active, sale_type)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
            (req.name, req.phone, req.email, req.region, req.address_details,
             req.tax_number, req.credit_limit, req.notes, req.discount_percent,
             req.discount_notes, req.active, (req.sale_type or "cash")),
        )
        row = cur.fetchone()
        final_code = _ensure_customer_code(cur, row["id"], req.code)
        cur.execute("UPDATE customers SET code=%s WHERE id=%s RETURNING *", (final_code, row["id"]))
        row = cur.fetchone()
        if req.branch_ids:
            for bid in set(req.branch_ids):
                cur.execute(
                    "INSERT INTO customer_branches (customer_id, branch_id, authorized_by) "
                    "VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                    (row["id"], bid, current_user.get("user_id")),
                )
        conn.commit()
        return dict(row)
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.put("/customers/v2/{customer_id}")
def update_customer_v2(customer_id: int, req: CustomerIn,
                       current_user=Depends(get_current_user)):
    _admin_only(current_user)
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        if is_platform_partner_customer_id(cur, customer_id):
            raise HTTPException(
                status_code=400,
                detail="Digital platform partner accounts are managed in Settings → Digital Platforms, not here.",
            )
        cur.execute("SELECT code FROM customers WHERE id=%s", (customer_id,))
        existing_code_row = cur.fetchone()
        if not existing_code_row:
            raise HTTPException(status_code=404, detail="Customer not found")
        if req.code not in (None, ""):
            final_code = _ensure_customer_code(cur, customer_id, req.code)
        else:
            final_code = (existing_code_row.get("code") or "").strip() or _default_customer_code(customer_id)
        cur.execute(
            """UPDATE customers SET code=%s, name=%s, phone=%s, email=%s, region=%s,
               address_details=%s, tax_number=%s, credit_limit=%s, notes=%s,
               discount_percent=%s, discount_notes=%s, active=%s,
               sale_type=COALESCE(%s, sale_type)
               WHERE id=%s RETURNING *""",
            (final_code, req.name, req.phone, req.email, req.region, req.address_details,
             req.tax_number, req.credit_limit, req.notes, req.discount_percent,
             req.discount_notes, req.active, req.sale_type, customer_id),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Customer not found")
        if req.branch_ids is not None:
            # Replace authorization set
            cur.execute("DELETE FROM customer_branches WHERE customer_id=%s", (customer_id,))
            for bid in set(req.branch_ids):
                cur.execute(
                    "INSERT INTO customer_branches (customer_id, branch_id, authorized_by) "
                    "VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                    (customer_id, bid, current_user.get("user_id")),
                )
        conn.commit()
        return dict(row)
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/customers/v2/{customer_id}/branches")
def list_customer_branches(customer_id: int, current_user=Depends(get_current_user)):
    """Branches authorized to sell on account for this customer (admin only)."""
    _admin_only(current_user)
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT cb.branch_id, b.name_en, b.name_ar FROM customer_branches cb "
        "JOIN branches b ON b.id = cb.branch_id WHERE cb.customer_id=%s",
        (customer_id,),
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


@router.delete("/customers/v2/{customer_id}")
def deactivate_customer(customer_id: int, current_user=Depends(get_current_user)):
    _admin_only(current_user)
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        if is_platform_partner_customer_id(cur, customer_id):
            raise HTTPException(
                status_code=400,
                detail="Digital platform partner accounts cannot be removed here — use Settings → Digital Platforms.",
            )
        cur.execute("UPDATE customers SET active=false WHERE id=%s", (customer_id,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Customer not found")
        conn.commit()
        return {"ok": True}
    except HTTPException:
        conn.rollback()
        raise
    finally:
        conn.close()


class CustPaymentIn(BaseModel):
    amount: float
    payment_method: Optional[str] = "cash"
    invoice_id: Optional[int] = None
    reference: Optional[str] = None
    notes: Optional[str] = None


@router.post("/customers/v2/{customer_id}/payments")
def record_customer_payment(customer_id: int, req: CustPaymentIn,
                            current_user=Depends(get_current_user)):
    if req.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be positive")
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT id FROM customers WHERE id=%s", (customer_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Customer not found")
        if req.invoice_id:
            cur.execute(
                "SELECT id, branch_id FROM invoices WHERE id=%s AND customer_id=%s",
                (req.invoice_id, customer_id),
            )
            inv = cur.fetchone()
            if not inv:
                raise HTTPException(status_code=400, detail="Invoice does not belong to this customer")
            # Non-admin can only collect against their branch's invoices
            if current_user.get("role") != "admin":
                if inv["branch_id"] != current_user.get("branch_id"):
                    raise HTTPException(status_code=403, detail="Cannot collect against another branch's invoice")
        else:
            # Untied payment is admin-only (cannot attribute to any branch)
            _admin_only(current_user)
        cur.execute(
            """INSERT INTO customer_payments
               (customer_id, invoice_id, amount, payment_method, reference, notes, recorded_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
            (customer_id, req.invoice_id, req.amount, req.payment_method,
             req.reference, req.notes, current_user.get("user_id")),
        )
        row = cur.fetchone()
        conn.commit()
        return dict(row)
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/customers/v2/{customer_id}/statement")
def customer_statement(customer_id: int, current_user=Depends(get_current_user)):
    """Chronological ledger: account-sale debits + payment credits + running balance.
    Branch-scoped for non-admins."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    # Branch-gate non-admins BEFORE returning any customer data
    if current_user.get("role") != "admin":
        ub = current_user.get("branch_id")
        if ub is None:
            conn.close()
            raise HTTPException(status_code=403, detail="No branch assigned")
        cur.execute(
            "SELECT 1 FROM customer_branches WHERE customer_id=%s AND branch_id=%s",
            (customer_id, ub),
        )
        if not cur.fetchone():
            conn.close()
            raise HTTPException(status_code=404, detail="Customer not found")
    cur.execute("SELECT * FROM customers WHERE id=%s", (customer_id,))
    cust = cur.fetchone()
    if not cust:
        conn.close()
        raise HTTPException(status_code=404, detail="Customer not found")
    if current_user.get("role") != "admin":
        ub = current_user.get("branch_id")
        cur.execute(
            """SELECT 'sale' AS kind, i.id AS ref_id, i.invoice_number AS reference,
                      i.created_at AS at, i.net_total AS debit, 0 AS credit, i.notes
               FROM invoices i
               WHERE i.customer_id=%s AND i.payment_method='account'
                     AND i.type!='return' AND i.branch_id=%s
               UNION ALL
               SELECT 'payment' AS kind, cp.id, cp.reference, cp.paid_at AS at,
                      0 AS debit, cp.amount AS credit, cp.notes
               FROM customer_payments cp
               JOIN invoices i ON cp.invoice_id = i.id
               WHERE cp.customer_id=%s AND i.branch_id=%s
               ORDER BY at ASC NULLS LAST""",
            (customer_id, ub, customer_id, ub),
        )
    else:
        cur.execute(
            """SELECT 'sale' AS kind, id AS ref_id, invoice_number AS reference,
                      created_at AS at, net_total AS debit, 0 AS credit, notes
               FROM invoices
               WHERE customer_id=%s AND payment_method='account' AND type!='return'
               UNION ALL
               SELECT 'payment' AS kind, id, reference, paid_at AS at,
                      0 AS debit, amount AS credit, notes
               FROM customer_payments
               WHERE customer_id=%s
               ORDER BY at ASC NULLS LAST""",
            (customer_id, customer_id),
        )
    txns = [dict(r) for r in cur.fetchall()]
    running = 0.0
    for t in txns:
        running += float(t["debit"] or 0) - float(t["credit"] or 0)
        t["balance"] = round(running, 2)
    conn.close()
    return {"customer": dict(cust), "transactions": txns, "balance": round(running, 2)}
