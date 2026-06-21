"""Purchasing & Suppliers module — suppliers, POs, supplier payments & statements."""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, datetime
import psycopg2.extras
from db import get_db_connection
from deps import get_current_user, get_active_branch_id
from inventory import log_movement
from stock_batches import add_batch_stock, sync_product_from_batches

router = APIRouter(prefix="/api", tags=["purchasing"])


def _admin_only(user):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admins only")


def _pharmacy_unit(public_price, sales_tax, discount_pct) -> float:
    """Pharmacy/cost price per unit — ibn sina: (public − tax) × (1 − disc%)."""
    pub = float(public_price or 0)
    tax = float(sales_tax or 0)
    base = (pub - tax) if tax > 0 else pub
    return base * (1 - float(discount_pct or 0) / 100)


def _line_ex_tax(qty, public_price, sales_tax, discount_pct, unit_cost=None) -> float:
    """Line total ex-VAT (الاجمالي بدون ضريبة)."""
    pub = float(public_price) if public_price is not None else float(unit_cost or 0)
    if public_price is None and unit_cost and not sales_tax:
        return int(qty) * float(unit_cost) * (1 - float(discount_pct or 0) / 100)
    unit = _pharmacy_unit(pub, sales_tax, discount_pct)
    return int(qty) * unit


def _line_tax_amount(qty, sales_tax) -> float:
    """Line sales tax total (ضريبة مبيعات)."""
    return int(qty) * float(sales_tax or 0)


def _line_gross(qty, public_price, sales_tax, discount_pct, unit_cost=None) -> float:
    return _line_ex_tax(qty, public_price, sales_tax, discount_pct, unit_cost) + _line_tax_amount(qty, sales_tax)


def _line_net(qty, unit_cost, discount_pct) -> float:
    """Legacy alias — prefer _line_ex_tax with public_price + sales_tax."""
    return int(qty) * float(unit_cost) * (1 - float(discount_pct or 0) / 100)


def _line_vat(net: float, vat_pct) -> float:
    return net * float(vat_pct or 0) / 100


def _resolve_line_pricing(it) -> tuple[float, float, float, float]:
    """Returns (public_price, pharmacy_unit_cost, sales_tax, retail_price)."""
    pub = float(it.public_price if it.public_price is not None else it.unit_cost or 0)
    tax = float(getattr(it, "sales_tax", None) or 0)
    disc = float(it.discount_pct or 0)
    pharmacy = _pharmacy_unit(pub, tax, disc)
    retail = float(it.public_price) if it.public_price is not None else pub
    return pub, pharmacy, tax, retail


def _eff_unit_cost(unit_cost, discount_pct, vat_pct):
    """Inventory cost = pharmacy price (discount already applied; tax is separate on invoice)."""
    return float(unit_cost or 0)


# ─── SUPPLIERS ──────────────────────────────────────────────────────────

class SupplierIn(BaseModel):
    name: str
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    region: Optional[str] = None
    address_details: Optional[str] = None
    tax_number: Optional[str] = None
    notes: Optional[str] = None
    active: bool = True


@router.get("/suppliers")
def list_suppliers(q: str = "", active_only: bool = True,
                   current_user=Depends(get_current_user)):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    where = []
    params: list = []
    if active_only:
        where.append("s.active = true")
    if q:
        where.append("(s.name ILIKE %s OR s.phone ILIKE %s OR s.contact_person ILIKE %s)")
        like = f"%{q}%"
        params += [like, like, like]
    # Aggregate totals: branch-scoped for non-admins so they can't see other branches' financials
    if current_user.get("role") != "admin":
        ub = current_user.get("branch_id")
        if ub is None:
            raise HTTPException(status_code=403, detail="No branch assigned")
        charged_sub = ("(SELECT COALESCE(SUM(total),0) FROM purchase_orders "
                       "WHERE supplier_id=s.id AND status='received' AND branch_id=%s)")
        # Payments tied to a PO in user's branch + payments not tied to any PO are hidden from non-admins
        paid_sub = ("(SELECT COALESCE(SUM(sp.amount),0) FROM supplier_payments sp "
                    "JOIN purchase_orders po ON sp.po_id=po.id "
                    "WHERE sp.supplier_id=s.id AND po.branch_id=%s)")
        select_aggs = f"{charged_sub} AS total_charged, {paid_sub} AS total_paid"
        sql = f"SELECT s.*, {select_aggs} FROM suppliers s"
        # Prepend two branch params before any q params
        params = [ub, ub] + params
    else:
        sql = """SELECT s.*,
                        COALESCE((SELECT SUM(total) FROM purchase_orders
                                  WHERE supplier_id=s.id AND status='received'), 0) AS total_charged,
                        COALESCE((SELECT SUM(amount) FROM supplier_payments
                                  WHERE supplier_id=s.id), 0) AS total_paid
                 FROM suppliers s"""
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY s.name ASC LIMIT 500"
    cur.execute(sql, params)
    rows = [dict(r) for r in cur.fetchall()]
    for r in rows:
        r["balance"] = float(r["total_charged"]) - float(r["total_paid"])
    conn.close()
    return rows


@router.post("/suppliers")
def create_supplier(req: SupplierIn, current_user=Depends(get_current_user)):
    _admin_only(current_user)
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            """INSERT INTO suppliers
               (name, contact_person, phone, email, address, region, address_details,
                tax_number, notes, active)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
            (req.name, req.contact_person, req.phone, req.email, req.address,
             req.region, req.address_details,
             req.tax_number, req.notes, req.active),
        )
        row = cur.fetchone()
        conn.commit()
        return dict(row)
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.put("/suppliers/{supplier_id}")
def update_supplier(supplier_id: int, req: SupplierIn,
                    current_user=Depends(get_current_user)):
    _admin_only(current_user)
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            """UPDATE suppliers SET name=%s, contact_person=%s, phone=%s, email=%s,
               address=%s, region=%s, address_details=%s,
               tax_number=%s, notes=%s, active=%s WHERE id=%s RETURNING *""",
            (req.name, req.contact_person, req.phone, req.email, req.address,
             req.region, req.address_details,
             req.tax_number, req.notes, req.active, supplier_id),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Supplier not found")
        conn.commit()
        return dict(row)
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.delete("/suppliers/{supplier_id}")
def delete_supplier(supplier_id: int, current_user=Depends(get_current_user)):
    _admin_only(current_user)
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE suppliers SET active=false WHERE id=%s", (supplier_id,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Supplier not found")
        conn.commit()
        return {"ok": True}
    except HTTPException:
        conn.rollback()
        raise
    finally:
        conn.close()


# ─── PURCHASE ORDERS ─────────────────────────────────────────────────────

class POItemIn(BaseModel):
    product_id: Optional[int] = None  # null = new product to be created on receive
    barcode: Optional[str] = None
    product_name_ar: Optional[str] = None
    product_name_en: Optional[str] = None
    quantity: int
    bonus_qty: int = 0
    unit_cost: float
    discount_pct: float = 0
    vat_pct: float = 0
    sales_tax: float = 0
    public_price: Optional[float] = None
    expiry_date: Optional[date] = None


class POIn(BaseModel):
    supplier_id: int
    branch_id: int
    supplier_invoice_number: Optional[str] = None
    supplier_invoice_date: Optional[date] = None
    discount: float = 0
    tax: float = 0
    notes: Optional[str] = None
    receive_immediately: bool = False
    items: List[POItemIn]


def _assert_po_branch_access(user, branch_id: int):
    if user.get("role") == "admin":
        return
    if user.get("branch_id") != branch_id:
        raise HTTPException(status_code=403, detail="Cannot manage POs for another branch")


def _inventory_unit_cost(paid_qty: int, bonus_qty: int, eff_cost: float) -> float:
    """Spread paid line cost across paid + bonus units (e.g. 10 paid + 2 free)."""
    total_units = paid_qty + bonus_qty
    if total_units <= 0:
        return eff_cost
    return (paid_qty * eff_cost) / total_units


def _receive_po_line(cur, it, po: dict, branch_id: int, current_user: dict) -> None:
    """Apply one PO line to branch stock (paid qty + bonus qty)."""
    eff_cost = _eff_unit_cost(it["unit_cost"], it.get("discount_pct"), it.get("vat_pct"))
    pub_price = it.get("public_price")
    new_price = float(pub_price) if pub_price not in (None, "") and float(pub_price) > 0 else None
    paid_qty = int(it["quantity"])
    bonus_qty = max(0, int(it.get("bonus_qty") or 0))
    stock_qty = paid_qty + bonus_qty
    if stock_qty <= 0:
        raise HTTPException(status_code=400, detail="Line quantity must be positive")
    inventory_cost = _inventory_unit_cost(paid_qty, bonus_qty, eff_cost)

    pid = it["product_id"]
    pack_size = 1
    if pid:
        cur.execute("SELECT id, stock, branch_id, pack_size FROM products WHERE id=%s FOR UPDATE", (pid,))
        p = cur.fetchone()
        if not p:
            raise HTTPException(status_code=400, detail=f"Product id {pid} no longer exists")
        if p["branch_id"] != branch_id:
            pid = None
        else:
            pack_size = max(1, int(p.get("pack_size") or 1))
    if not pid and it["barcode"]:
        cur.execute(
            "SELECT id, stock, pack_size FROM products WHERE barcode=%s AND branch_id=%s FOR UPDATE",
            (it["barcode"], branch_id),
        )
        p = cur.fetchone()
        if p:
            pid = p["id"]
            pack_size = max(1, int(p.get("pack_size") or 1))
    if not pid:
        cur.execute(
            """INSERT INTO products
               (barcode, name_ar, name_en, category, unit, price, cost,
                stock, min_stock, expiry_date, branch_id, active)
               VALUES (%s,%s,%s,'',' ',%s,%s,0,5,%s,%s,true)
               ON CONFLICT (barcode, branch_id) DO NOTHING
               RETURNING id, stock""",
            (it["barcode"], it["product_name_ar"] or it["product_name_en"],
             it["product_name_en"] or it["product_name_ar"],
             new_price if new_price is not None else inventory_cost, inventory_cost,
             it["expiry_date"], branch_id),
        )
        p = cur.fetchone()
        if not p and it["barcode"]:
            cur.execute(
                "SELECT id, stock, pack_size FROM products WHERE barcode=%s AND branch_id=%s FOR UPDATE",
                (it["barcode"], branch_id),
            )
            p = cur.fetchone()
        if not p:
            raise HTTPException(
                status_code=400,
                detail=f"Could not create/find product for '{it.get('product_name_en') or it.get('barcode')}' — barcode required for new items",
            )
        pid = p["id"]
        pack_size = max(1, int(p.get("pack_size") or 1))

    # PO qty is in packs/boxes — stock moves 1:1 with entered quantity.
    outer_qty = stock_qty
    stock_qty_sub = outer_qty
    add_batch_stock(cur, pid, branch_id, stock_qty_sub, it["expiry_date"])
    new_stock = sync_product_from_batches(cur, pid)
    sets = ["cost=%s"]
    params: list = [inventory_cost]
    if new_price is not None:
        sets.append("price=%s")
        params.append(new_price)
        if pack_size > 1:
            sets.append("sub_price=%s")
            params.append(round(new_price / pack_size, 2))
    params.append(pid)
    cur.execute(f"UPDATE products SET {', '.join(sets)} WHERE id=%s", params)
    cur.execute("UPDATE purchase_order_items SET product_id=%s WHERE id=%s", (pid, it["id"]))
    log_movement(
        cur, pid, branch_id, "purchase",
        stock_qty_sub, new_stock,
        reference_type="po", reference_id=po["id"],
        reason=f"PO {po['po_number']} received"
        + (f" ({paid_qty}+{bonus_qty} bonus)" if bonus_qty else ""),
        user_id=current_user.get("user_id"),
    )


@router.post("/purchase-orders")
def create_po(req: POIn, current_user=Depends(get_current_user)):
    if not req.items:
        raise HTTPException(status_code=400, detail="At least one item required")
    if req.discount < 0 or req.tax < 0:
        raise HTTPException(status_code=400, detail="Discount and tax cannot be negative")
    _assert_po_branch_access(current_user, req.branch_id)
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT id FROM suppliers WHERE id=%s AND active=true", (req.supplier_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=400, detail="Invalid supplier")
        cur.execute("SELECT id FROM branches WHERE id=%s", (req.branch_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=400, detail="Invalid branch")

        for it in req.items:
            if it.quantity <= 0:
                raise HTTPException(status_code=400, detail="Quantity must be positive")
            if it.bonus_qty < 0:
                raise HTTPException(status_code=400, detail="Bonus quantity cannot be negative")
            if it.unit_cost < 0:
                raise HTTPException(status_code=400, detail="Unit cost cannot be negative")
            if it.discount_pct < 0 or it.discount_pct > 100:
                raise HTTPException(status_code=400, detail="Discount % must be between 0 and 100")
            if it.vat_pct < 0:
                raise HTTPException(status_code=400, detail="VAT % cannot be negative")
            if it.public_price is not None and it.public_price < 0:
                raise HTTPException(status_code=400, detail="Public price cannot be negative")
        # Supplier consistency: if any selected product has a preferred supplier set,
        # it must match the PO's supplier (NULL = unassigned, allowed). Prevents
        # the replenishment flow (or any UI) from sending mixed-supplier POs.
        product_ids = [int(i.product_id) for i in req.items if i.product_id]
        if product_ids:
            cur.execute(
                "SELECT id, supplier_id FROM products WHERE id = ANY(%s)",
                (product_ids,),
            )
            mismatched = [int(r["id"]) for r in cur.fetchall()
                          if r["supplier_id"] not in (None, req.supplier_id)]
            if mismatched:
                raise HTTPException(
                    status_code=400,
                    detail=f"Items belong to a different supplier (ids: {mismatched[:10]})",
                )
        subtotal_net = sum(
            _line_ex_tax(i.quantity, i.public_price, i.sales_tax, i.discount_pct, i.unit_cost)
            for i in req.items
        )
        total_vat = sum(_line_tax_amount(i.quantity, i.sales_tax) for i in req.items)
        total = subtotal_net - req.discount + total_vat + req.tax
        if total < 0:
            raise HTTPException(status_code=400, detail="Discount cannot exceed invoice total")

        cur.execute("SELECT nextval('purchase_order_seq') AS n")
        seq_n = cur.fetchone()["n"]
        po_number = f"PO-{date.today().strftime('%Y%m%d')}-{int(seq_n):04d}"

        cur.execute(
            """INSERT INTO purchase_orders
               (po_number, supplier_id, branch_id, status,
                supplier_invoice_number, supplier_invoice_date,
                subtotal, discount, tax, total, notes, created_by)
               VALUES (%s,%s,%s,'draft',%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            (po_number, req.supplier_id, req.branch_id,
             req.supplier_invoice_number, req.supplier_invoice_date,
             subtotal_net, req.discount, total_vat + req.tax, total, req.notes,
             current_user.get("user_id")),
        )
        po_id = cur.fetchone()["id"]

        for it in req.items:
            if it.quantity <= 0:
                raise HTTPException(status_code=400, detail="Quantity must be positive")
            if it.unit_cost < 0:
                raise HTTPException(status_code=400, detail="Unit cost cannot be negative")
            # Snapshot product names from product if product_id given
            pname_ar = it.product_name_ar
            pname_en = it.product_name_en
            barcode = it.barcode
            if it.product_id:
                cur.execute("SELECT name_ar, name_en, barcode FROM products WHERE id=%s",
                            (it.product_id,))
                p = cur.fetchone()
                if not p:
                    raise HTTPException(status_code=404, detail=f"Product {it.product_id} not found")
                pname_ar = pname_ar or p["name_ar"]
                pname_en = pname_en or p["name_en"]
                barcode = barcode or p["barcode"]
            if not pname_en:
                raise HTTPException(status_code=400, detail="Product name required")
            pub, pharmacy, stax, retail = _resolve_line_pricing(it)
            line_total = _line_gross(it.quantity, pub, stax, it.discount_pct, pharmacy)
            cur.execute(
                """INSERT INTO purchase_order_items
                   (po_id, product_id, barcode, product_name_ar, product_name_en,
                    quantity, bonus_qty, unit_cost, discount_pct, vat_pct, sales_tax, public_price,
                    expiry_date, total)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (po_id, it.product_id, barcode, pname_ar, pname_en,
                 it.quantity, it.bonus_qty, pharmacy, it.discount_pct, 0, stax,
                 retail, it.expiry_date, line_total),
            )

        po_row = {"id": po_id, "po_number": po_number, "branch_id": req.branch_id}
        status = "draft"
        if req.receive_immediately:
            cur.execute("SELECT * FROM purchase_order_items WHERE po_id=%s", (po_id,))
            for row in cur.fetchall():
                _receive_po_line(cur, row, po_row, req.branch_id, current_user)
            cur.execute(
                """UPDATE purchase_orders
                   SET status='received', received_by=%s, received_at=NOW()
                   WHERE id=%s""",
                (current_user.get("user_id"), po_id),
            )
            status = "received"

        conn.commit()
        return {
            "ok": True,
            "po_id": po_id,
            "po_number": po_number,
            "total": total,
            "status": status,
        }
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/purchase-orders")
def list_pos(status: Optional[str] = None, supplier_id: Optional[int] = None,
             current_user=Depends(get_current_user),
             active_branch=Depends(get_active_branch_id)):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    where = []
    params: list = []
    if status:
        where.append("po.status = %s")
        params.append(status)
    if supplier_id:
        where.append("po.supplier_id = %s")
        params.append(supplier_id)
    if current_user.get("role") != "admin":
        ub = current_user.get("branch_id")
        if ub is None:
            raise HTTPException(status_code=403, detail="No branch assigned")
        where.append("po.branch_id = %s")
        params.append(ub)
    elif active_branch is not None:
        where.append("po.branch_id = %s")
        params.append(active_branch)
    sql = """SELECT po.*, s.name AS supplier_name,
                    b.name_en AS branch_name_en, b.name_ar AS branch_name_ar
             FROM purchase_orders po
             LEFT JOIN suppliers s ON po.supplier_id = s.id
             LEFT JOIN branches b ON po.branch_id = b.id"""
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY po.created_at DESC LIMIT 200"
    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]


@router.get("/purchase-orders/{po_id:int}")
def get_po(po_id: int, current_user=Depends(get_current_user)):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """SELECT po.*, s.name AS supplier_name,
                  b.name_en AS branch_name_en, b.name_ar AS branch_name_ar
           FROM purchase_orders po
           LEFT JOIN suppliers s ON po.supplier_id = s.id
           LEFT JOIN branches b ON po.branch_id = b.id
           WHERE po.id=%s""",
        (po_id,),
    )
    po = cur.fetchone()
    if not po:
        conn.close()
        raise HTTPException(status_code=404, detail="PO not found")
    if current_user.get("role") != "admin":
        if current_user.get("branch_id") != po["branch_id"]:
            conn.close()
            raise HTTPException(status_code=403, detail="Not accessible")
    cur.execute("SELECT * FROM purchase_order_items WHERE po_id=%s ORDER BY id", (po_id,))
    items = cur.fetchall()
    conn.close()
    out = dict(po)
    out["items"] = [dict(i) for i in items]
    return out


@router.post("/purchase-orders/{po_id:int}/receive")
def receive_po(po_id: int, current_user=Depends(get_current_user)):
    """Receive PO: add stock to branch, log purchase movements,
    update product cost + expiry_date when provided."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT * FROM purchase_orders WHERE id=%s FOR UPDATE", (po_id,))
        po = cur.fetchone()
        if not po:
            raise HTTPException(status_code=404, detail="PO not found")
        if po["status"] != "draft":
            raise HTTPException(status_code=400, detail=f"Cannot receive PO in status '{po['status']}'")
        _assert_po_branch_access(current_user, po["branch_id"])
        branch_id = po["branch_id"]

        cur.execute("SELECT * FROM purchase_order_items WHERE po_id=%s", (po_id,))
        items = cur.fetchall()
        for it in items:
            _receive_po_line(cur, it, po, branch_id, current_user)

        cur.execute(
            """UPDATE purchase_orders
               SET status='received', received_by=%s, received_at=NOW()
               WHERE id=%s""",
            (current_user.get("user_id"), po_id),
        )
        conn.commit()
        return {"ok": True, "status": "received"}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.post("/purchase-orders/{po_id:int}/cancel")
def cancel_po(po_id: int, current_user=Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT * FROM purchase_orders WHERE id=%s FOR UPDATE", (po_id,))
        po = cur.fetchone()
        if not po:
            raise HTTPException(status_code=404, detail="PO not found")
        if po["status"] != "draft":
            raise HTTPException(status_code=400, detail=f"Cannot cancel PO in status '{po['status']}'")
        cur.execute(
            "UPDATE purchase_orders SET status='cancelled', cancelled_at=NOW() WHERE id=%s",
            (po_id,),
        )
        conn.commit()
        return {"ok": True, "status": "cancelled"}
    except HTTPException:
        conn.rollback()
        raise
    finally:
        conn.close()


# ─── SUPPLIER PAYMENTS & STATEMENT ────────────────────────────────────────

class PaymentIn(BaseModel):
    amount: float
    payment_method: Optional[str] = "cash"
    po_id: Optional[int] = None
    reference: Optional[str] = None
    notes: Optional[str] = None


@router.post("/suppliers/{supplier_id}/payments")
def record_payment(supplier_id: int, req: PaymentIn,
                   current_user=Depends(get_current_user)):
    _admin_only(current_user)
    if req.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be positive")
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT id FROM suppliers WHERE id=%s", (supplier_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Supplier not found")
        if req.po_id:
            cur.execute("SELECT id FROM purchase_orders WHERE id=%s AND supplier_id=%s",
                        (req.po_id, supplier_id))
            if not cur.fetchone():
                raise HTTPException(status_code=400, detail="PO does not belong to this supplier")
        cur.execute(
            """INSERT INTO supplier_payments
               (supplier_id, po_id, amount, payment_method, reference, notes, recorded_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
            (supplier_id, req.po_id, req.amount, req.payment_method,
             req.reference, req.notes, current_user.get("user_id")),
        )
        row = cur.fetchone()
        conn.commit()
        return dict(row)
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/suppliers/{supplier_id}/statement")
def supplier_statement(supplier_id: int,
                       current_user=Depends(get_current_user)):
    """Combined chronological ledger of PO charges and payments with running balance.

    Branch-scoped for non-admins: only rows tied to their branch are visible. Payments
    not tied to any PO are admin-only (cannot be branch-attributed).
    """
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM suppliers WHERE id=%s", (supplier_id,))
    sup = cur.fetchone()
    if not sup:
        conn.close()
        raise HTTPException(status_code=404, detail="Supplier not found")
    if current_user.get("role") != "admin":
        ub = current_user.get("branch_id")
        if ub is None:
            conn.close()
            raise HTTPException(status_code=403, detail="No branch assigned")
        cur.execute(
            """SELECT 'po' AS kind, po.id AS ref_id, po.po_number AS reference,
                      po.received_at AS at, po.total AS debit, 0 AS credit, po.notes
               FROM purchase_orders po
               WHERE po.supplier_id=%s AND po.status='received' AND po.branch_id=%s
               UNION ALL
               SELECT 'payment' AS kind, sp.id, sp.reference, sp.paid_at AS at,
                      0 AS debit, sp.amount AS credit, sp.notes
               FROM supplier_payments sp
               JOIN purchase_orders po ON sp.po_id = po.id
               WHERE sp.supplier_id=%s AND po.branch_id=%s
               ORDER BY at ASC NULLS LAST""",
            (supplier_id, ub, supplier_id, ub),
        )
    else:
        cur.execute(
            """SELECT 'po' AS kind, id AS ref_id, po_number AS reference,
                      received_at AS at, total AS debit, 0 AS credit, notes
               FROM purchase_orders
               WHERE supplier_id=%s AND status='received'
               UNION ALL
               SELECT 'payment' AS kind, id, reference, paid_at AS at,
                      0 AS debit, amount AS credit, notes
               FROM supplier_payments
               WHERE supplier_id=%s
               ORDER BY at ASC NULLS LAST""",
            (supplier_id, supplier_id),
        )
    txns = [dict(r) for r in cur.fetchall()]
    running = 0.0
    for t in txns:
        running += float(t["debit"] or 0) - float(t["credit"] or 0)
        t["balance"] = round(running, 2)
    conn.close()
    return {
        "supplier": dict(sup),
        "transactions": txns,
        "balance": round(running, 2),
    }


# ─── REPLENISHMENT (auto-PO from low-stock) ──────────────────────────────

@router.get("/purchase-orders/replenishment")
def replenishment_list(
    branch_id: Optional[int] = None,
    supplier_id: Optional[int] = None,
    only_zero: bool = False,
    include_all: bool = False,
    q: str = "",
    current_user=Depends(get_current_user),
    active_branch=Depends(get_active_branch_id),
):
    """List items that need replenishment (stock <= min_stock).
    When include_all is set, every active product is returned and each row
    carries needs_replenish so the UI can pre-select only the low ones.
    Pass q to search by name/barcode without loading the full catalog.
    Returns [] unless include_all, q, or only_zero is set (avoids slow full loads).
    Non-admins are constrained to their own branch."""
    q = (q or "").strip()
    if not include_all and not q and not only_zero:
        return []
    if current_user.get("role") != "admin":
        ub = current_user.get("branch_id")
        if ub is None:
            raise HTTPException(status_code=403, detail="No branch assigned")
        if branch_id is not None and branch_id != ub:
            raise HTTPException(status_code=403, detail="Cross-branch access denied")
        branch_id = ub
    eff_branch = branch_id

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    where = ["p.active = true"]
    params: list = []
    if eff_branch is not None:
        where.append("p.branch_id = %s")
        params.append(eff_branch)
    if supplier_id is not None:
        where.append("p.supplier_id = %s")
        params.append(supplier_id)
    if q:
        where.append(
            "(p.name_ar ILIKE %s OR p.name_en ILIKE %s OR p.barcode ILIKE %s OR p.international_barcode ILIKE %s)"
        )
        s = f"%{q}%"
        params += [s, s, s, s]
    if only_zero:
        where.append("p.stock <= 0")
    elif not include_all and not q:
        where.append("p.stock <= p.min_stock")

    sql = (
        "SELECT p.id, p.barcode, p.name_ar, p.name_en, p.unit, p.sub_unit, p.pack_size, "
        "       p.stock, p.min_stock, p.cost, p.branch_id, p.supplier_id, "
        "       s.name AS supplier_name, "
        "       b.name_en AS branch_name_en, b.name_ar AS branch_name_ar "
        "FROM products p "
        "LEFT JOIN suppliers s ON p.supplier_id = s.id "
        "LEFT JOIN branches  b ON p.branch_id  = b.id "
        "WHERE " + " AND ".join(where) +
        " ORDER BY (p.stock <= 0) DESC, COALESCE(s.name, 'zz') ASC, p.name_en ASC"
    )
    cur.execute(sql, params)
    rows = []
    for r in cur.fetchall():
        d = dict(r)
        pack = max(1, int(d.get("pack_size") or 1))
        # Suggested order qty = bring stock back to 2× min_stock, with a sane floor.
        target = max(int(d.get("min_stock") or 0) * 2, int(d.get("min_stock") or 0) + 1, 1)
        suggested_sub = max(target - int(d.get("stock") or 0), 1)
        # PO quantities are expressed in the main unit (box). Convert the sub-unit gap to
        # whole boxes (round up so we never under-order).
        d["suggested_quantity"] = -(-suggested_sub // pack) if pack > 1 else suggested_sub
        d["needs_replenish"] = int(d.get("stock") or 0) <= int(d.get("min_stock") or 0)
        # PO is ordered in the main unit; label accordingly.
        d["unit_label"] = d.get("unit") or "unit"
        rows.append(d)
    conn.close()
    return rows


class ReplenishItemIn(BaseModel):
    product_id: int
    quantity: int
    unit_cost: float = 0


class ReplenishExportIn(BaseModel):
    supplier_id: Optional[int] = None  # for header only
    branch_id: Optional[int] = None    # for header only
    notes: Optional[str] = None
    items: List[ReplenishItemIn]


@router.post("/purchase-orders/replenishment/export")
def replenishment_export(req: ReplenishExportIn,
                         current_user=Depends(get_current_user)):
    """Build a supplier-ready Excel order sheet from the chosen replenishment items."""
    if not req.items:
        raise HTTPException(status_code=400, detail="No items selected")

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    def _safe(v):
        """Neutralize Excel/CSV formula injection on user/DB strings.
        Order matters: strip control chars and leading whitespace FIRST, then check
        the first visible character so an attacker can't smuggle '\\x01=SUM(...)'."""
        if v is None:
            return ""
        s = str(v)
        # Drop control chars (allow tab/newline) — openpyxl rejects most anyway.
        s = "".join(ch for ch in s if ch in ("\n", "\t") or ord(ch) >= 32)
        s = s.lstrip(" \t\n")
        if s and s[0] in ("=", "+", "-", "@"):
            s = "'" + s
        return s

    # Fail closed for non-admins without a branch.
    if current_user.get("role") != "admin":
        ub = current_user.get("branch_id")
        if ub is None:
            raise HTTPException(status_code=403, detail="No branch assigned")
        if req.branch_id is not None and req.branch_id != ub:
            raise HTTPException(status_code=403, detail="Cross-branch access denied")

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # Snapshot pharmacy + supplier + branch info for the sheet header.
        cur.execute("SELECT name_ar, name_en, phone, address_ar, address_en, tax_id FROM pharmacy_profile WHERE id=1")
        prof = cur.fetchone() or {}

        sup = None
        if req.supplier_id:
            cur.execute("SELECT id, name, contact_person, phone, email, address FROM suppliers WHERE id=%s",
                        (req.supplier_id,))
            sup = cur.fetchone()
            if not sup:
                raise HTTPException(status_code=400, detail="Invalid supplier")

        branch = None
        if req.branch_id:
            cur.execute("SELECT name_en, name_ar FROM branches WHERE id=%s", (req.branch_id,))
            branch = cur.fetchone()

        # Fetch each product line with safe parameterized lookup.
        ids = [int(i.product_id) for i in req.items]
        cur.execute(
            "SELECT id, barcode, name_ar, name_en, unit, sub_unit, pack_size, stock, min_stock, cost, branch_id, supplier_id "
            "FROM products WHERE id = ANY(%s)",
            (ids,),
        )
        by_id = {int(r["id"]): dict(r) for r in cur.fetchall()}
        # Every requested product must exist (otherwise silent drops).
        for pid in ids:
            if pid not in by_id:
                raise HTTPException(status_code=404, detail=f"Product {pid} not found")

        # Non-admins: every line must belong to their branch.
        if current_user.get("role") != "admin":
            ub = current_user.get("branch_id")
            for pid, p in by_id.items():
                if p["branch_id"] != ub:
                    raise HTTPException(status_code=403, detail="Cross-branch product in selection")

        # If supplier was specified, all selected products must belong to that supplier
        # so the supplier-addressed Excel is operationally correct.
        if req.supplier_id:
            mismatched = [int(pid) for pid, p in by_id.items()
                          if p.get("supplier_id") not in (None, req.supplier_id)]
            if mismatched:
                raise HTTPException(
                    status_code=400,
                    detail=f"Some selected items belong to a different supplier (ids: {mismatched[:10]})",
                )
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    bold = Font(bold=True, size=12)
    title = Font(bold=True, size=16, color="FFFFFF")
    fill = PatternFill("solid", fgColor="0F766E")
    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    center = Alignment(horizontal="center", vertical="center")

    # Title row
    ws.merge_cells("A1:G1")
    ws["A1"] = "Purchase Order / أمر شراء"
    ws["A1"].font = title
    ws["A1"].fill = fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    row = 3
    def kv(label: str, value):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=_safe(value) or "—")
        row += 1

    kv("Pharmacy", prof.get("name_en") or prof.get("name_ar"))
    if prof.get("phone"): kv("Pharmacy phone", prof.get("phone"))
    if prof.get("tax_id"): kv("Tax ID", prof.get("tax_id"))
    if branch: kv("Branch", branch.get("name_en") or branch.get("name_ar"))
    kv("Date", date.today().isoformat())

    if sup:
        row += 1
        ws.cell(row=row, column=1, value="── Supplier ──").font = bold
        row += 1
        kv("Supplier", sup.get("name"))
        if sup.get("contact_person"): kv("Contact", sup.get("contact_person"))
        if sup.get("phone"):          kv("Phone",   sup.get("phone"))
        if sup.get("email"):          kv("Email",   sup.get("email"))
        if sup.get("address"):        kv("Address", sup.get("address"))

    if req.notes:
        row += 1
        ws.cell(row=row, column=1, value="Notes").font = bold
        ws.cell(row=row, column=2, value=_safe(req.notes))
        row += 1

    row += 2
    headers = ["#", "Barcode", "Product (EN)", "Product (AR)", "Unit",
               "Order Qty", "Unit Cost", "Line Total"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = center
        cell.border = border
    row += 1

    grand_total = 0.0
    for idx, it in enumerate(req.items, start=1):
        if it.quantity <= 0:
            raise HTTPException(status_code=400, detail="Quantity must be positive")
        if it.unit_cost < 0:
            raise HTTPException(status_code=400, detail="Unit cost cannot be negative")
        p = by_id.get(int(it.product_id))
        if not p:
            raise HTTPException(status_code=404, detail=f"Product {it.product_id} not found")
        # PO/order quantities are in the main unit (box); unit_cost is per box.
        unit_label = p.get("unit") or "unit"
        line_total = round(it.quantity * it.unit_cost, 2)
        grand_total += line_total
        values = [idx, _safe(p.get("barcode")), _safe(p.get("name_en")), _safe(p.get("name_ar")),
                  _safe(unit_label), it.quantity, round(it.unit_cost, 2), line_total]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = border
            if c in (1, 5, 6):
                cell.alignment = center
        row += 1

    # Grand total row
    ws.cell(row=row, column=7, value="Total").font = bold
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
    tcell = ws.cell(row=row, column=8, value=round(grand_total, 2))
    tcell.font = bold
    tcell.fill = PatternFill("solid", fgColor="ECFDF5")
    tcell.border = border

    # Column widths
    widths = [5, 18, 32, 32, 10, 12, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Filename: ASCII-safe, no path/quote chars so the Content-Disposition header is safe.
    raw = (sup["name"] if sup else "all")
    sup_part = "".join(c if (c.isalnum() or c in ("-", "_")) else "_" for c in str(raw))[:40] or "all"
    filename = f"PO_{sup_part}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
