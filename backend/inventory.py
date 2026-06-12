"""Inventory & Stock module — items, manual adjustments, movement ledger."""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, datetime
import io
import math
import psycopg2.extras
from db import get_db_connection
from deps import get_current_user, get_active_branch_id, requires_feature
from stock_batches import (
    add_batch_stock,
    deduct_stock_fefo,
    list_batches,
    set_batch_quantity,
    set_product_stock_absolute,
    sync_product_from_batches,
)

router = APIRouter(prefix="/api/inventory", tags=["inventory"])


def _xlsx_safe(v):
    if v is None:
        return ""
    s = str(v)
    s = "".join(ch for ch in s if ch in ("\n", "\t") or ord(ch) >= 32)
    s = s.lstrip(" \t\n")
    if s and s[0] in ("=", "+", "-", "@"):
        s = "'" + s
    return s


def _xlsx_response(headers: list, rows: list, filename: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.append([_xlsx_safe(h) for h in headers])
    head_fill = PatternFill("solid", fgColor="1F8A4C")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([_xlsx_safe(c) for c in row])
    for col in ws.columns:
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max(width, 10), 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _assert_branch_access(user, product_branch_id):
    """Non-admin users may only act on products in their own branch."""
    if user.get("role") == "admin":
        return
    user_branch = user.get("branch_id")
    if user_branch is None:
        raise HTTPException(status_code=403, detail="Cross-branch access denied")
    if product_branch_id is not None and product_branch_id != user_branch:
        raise HTTPException(status_code=403, detail="Cross-branch access denied")


def log_movement(cur, product_id: int, branch_id, movement_type: str,
                 quantity: int, balance_after: int,
                 reference_type: Optional[str] = None,
                 reference_id: Optional[int] = None,
                 reason: Optional[str] = None,
                 user_id: Optional[int] = None):
    """Insert a stock movement row. Caller manages transaction."""
    cur.execute(
        """INSERT INTO stock_movements
           (product_id, branch_id, movement_type, quantity, balance_after,
            reference_type, reference_id, reason, user_id)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (product_id, branch_id, movement_type, quantity, balance_after,
         reference_type, reference_id, reason, user_id),
    )


def _variance_split(delta_sub: int, pack_size: int) -> tuple[float, float]:
    """Return (whole_boxes, fractional_box) for a sub-unit delta."""
    if pack_size <= 1:
        return float(delta_sub), 0.0
    total_boxes = delta_sub / pack_size
    if total_boxes == 0:
        return 0.0, 0.0
    sign = 1 if total_boxes > 0 else -1
    abs_boxes = abs(total_boxes)
    whole = math.trunc(abs_boxes) * sign
    frac = total_boxes - whole
    return whole, frac


def _format_stocktake_line(row: dict) -> dict:
    pack = row.get("pack_size") or 1
    if pack <= 1:
        pack = 1
    delta = int(row["delta"])
    major, sub_frac = _variance_split(delta, pack)
    out = {
        "product_id": row["product_id"],
        "name_en": row.get("name_en"),
        "name_ar": row.get("name_ar"),
        "barcode": row.get("barcode"),
        "category": row.get("new_category") or row.get("category"),
        "pack_size": pack,
        "unit": row.get("unit") or "box",
        "sub_unit": row.get("sub_unit"),
        "old_stock": int(row["old_stock"]),
        "new_stock": int(row["new_stock"]),
        "delta": delta,
        "variance_major": major,
        "variance_sub_fraction": sub_frac,
        "old_category": row.get("old_category"),
        "new_category": row.get("new_category"),
        "old_expiry": row["old_expiry"].isoformat() if row.get("old_expiry") else None,
        "new_expiry": row["new_expiry"].isoformat() if row.get("new_expiry") else None,
    }
    return out


def _build_stocktake_report(cur, run_id: int) -> dict:
    cur.execute(
        """
        SELECT r.id AS run_id, r.branch_id, r.note, r.created_at,
               b.name_en AS branch_name_en, b.name_ar AS branch_name_ar,
               u.name_en AS user_name_en, u.name_ar AS user_name_ar
        FROM stocktake_runs r
        JOIN branches b ON b.id = r.branch_id
        LEFT JOIN users u ON u.id = r.user_id
        WHERE r.id = %s
        """,
        (run_id,),
    )
    run = cur.fetchone()
    if not run:
        raise HTTPException(status_code=404, detail="Stocktake run not found")
    run = dict(run)
    created = run.get("created_at")
    if created and hasattr(created, "isoformat"):
        run["created_at"] = created.isoformat()

    cur.execute(
        """
        SELECT l.*,
               p.name_en, p.name_ar, p.barcode, p.category,
               p.pack_size, p.unit, p.sub_unit
        FROM stocktake_lines l
        JOIN products p ON p.id = l.product_id
        WHERE l.run_id = %s
        ORDER BY ABS(l.delta) DESC, p.name_en ASC, p.name_ar ASC
        """,
        (run_id,),
    )
    lines = [_format_stocktake_line(dict(r)) for r in cur.fetchall()]
    shortages = [ln for ln in lines if ln["delta"] < 0]
    increases = [ln for ln in lines if ln["delta"] > 0]
    other = [ln for ln in lines if ln["delta"] == 0]

    return {
        **run,
        "summary": {
            "total_lines": len(lines),
            "shortages_count": len(shortages),
            "increases_count": len(increases),
            "other_count": len(other),
            "shortage_units": sum(ln["delta"] for ln in shortages),
            "increase_units": sum(ln["delta"] for ln in increases),
        },
        "shortages": shortages,
        "increases": increases,
        "other_changes": other,
        "lines": lines,
    }


# ─── PRODUCT CRUD (extends /api/products) ─────────────────────────────────

class ProductUpdate(BaseModel):
    barcode: Optional[str] = None
    international_barcode: Optional[str] = None
    name_ar: Optional[str] = None
    name_en: Optional[str] = None
    category: Optional[str] = None
    unit: Optional[str] = None
    price: Optional[float] = None
    cost: Optional[float] = None
    min_stock: Optional[int] = None
    expiry_date: Optional[date] = None
    active: Optional[bool] = None
    pack_size: Optional[int] = None
    sub_unit: Optional[str] = None
    sub_price: Optional[float] = None




ALLOWED_UPDATE_FIELDS = {"barcode", "international_barcode", "name_ar", "name_en", "category", "unit",
                         "price", "cost", "min_stock", "expiry_date", "active",
                         "pack_size", "sub_unit", "sub_price"}


@router.put("/products/{product_id}")
def update_product(product_id: int, req: ProductUpdate,
                   current_user=Depends(get_current_user)):
    fields = {k: v for k, v in req.model_dump(exclude_unset=True).items()
              if k in ALLOWED_UPDATE_FIELDS}
    if not fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT branch_id FROM products WHERE id=%s", (product_id,))
        existing = cur.fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Product not found")
        _assert_branch_access(current_user, existing["branch_id"])
        sets = ", ".join(f"{k}=%s" for k in fields.keys())
        values = list(fields.values()) + [product_id]
        cur.execute(f"UPDATE products SET {sets} WHERE id=%s RETURNING *", values)
        product = cur.fetchone()
        conn.commit()
        return dict(product)
    finally:
        conn.close()


@router.delete("/products/{product_id}")
def delete_product(product_id: int,
                   current_user=Depends(get_current_user)):
    """Soft delete — sets active=false. Admin only."""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT branch_id FROM products WHERE id=%s", (product_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Product not found")
        _assert_branch_access(current_user, row["branch_id"])
        cur.execute("UPDATE products SET active=false WHERE id=%s", (product_id,))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


# ─── STOCK BATCHES (multiple expiry per product) ───────────────────────────

class BatchCreate(BaseModel):
    expiry_date: Optional[date] = None
    quantity: int


class BatchUpdate(BaseModel):
    expiry_date: Optional[date] = None
    quantity: Optional[int] = None


@router.get("/products/{product_id}/batches")
def get_product_batches(product_id: int, current_user=Depends(get_current_user)):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT branch_id FROM products WHERE id=%s", (product_id,))
        p = cur.fetchone()
        if not p:
            raise HTTPException(status_code=404, detail="Product not found")
        _assert_branch_access(current_user, p["branch_id"])
        rows = list_batches(cur, product_id)
        return rows
    finally:
        conn.close()


@router.post("/products/{product_id}/batches")
def create_product_batch(
    product_id: int,
    req: BatchCreate,
    current_user=Depends(get_current_user),
):
    if req.quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be positive")
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT branch_id FROM products WHERE id=%s FOR UPDATE", (product_id,))
        p = cur.fetchone()
        if not p:
            raise HTTPException(status_code=404, detail="Product not found")
        _assert_branch_access(current_user, p["branch_id"])
        add_batch_stock(cur, product_id, p["branch_id"], req.quantity, req.expiry_date)
        new_stock = sync_product_from_batches(cur, product_id)
        log_movement(
            cur, product_id, p["branch_id"], "adjustment",
            req.quantity, new_stock,
            reference_type="batch", reason="Expiry lot added",
            user_id=current_user.get("user_id"),
        )
        conn.commit()
        return {"ok": True, "batches": list_batches(cur, product_id)}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.put("/batches/{batch_id}")
def update_product_batch(
    batch_id: int,
    req: BatchUpdate,
    current_user=Depends(get_current_user),
):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            """SELECT pb.*, p.branch_id FROM product_batches pb
               JOIN products p ON p.id = pb.product_id WHERE pb.id=%s FOR UPDATE""",
            (batch_id,),
        )
        batch = cur.fetchone()
        if not batch:
            raise HTTPException(status_code=404, detail="Batch not found")
        _assert_branch_access(current_user, batch["branch_id"])
        qty = req.quantity if req.quantity is not None else int(batch["quantity"])
        expiry_set = req.expiry_date is not None
        result = set_batch_quantity(
            cur, batch_id, qty, req.expiry_date, expiry_set=expiry_set,
        )
        delta = qty - int(batch["quantity"])
        if delta != 0:
            log_movement(
                cur, batch["product_id"], batch["branch_id"], "adjustment",
                delta, result["new_stock"],
                reference_type="batch", reason="Expiry lot updated",
                user_id=current_user.get("user_id"),
            )
        conn.commit()
        return {"ok": True, "batches": list_batches(cur, batch["product_id"])}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.delete("/batches/{batch_id}")
def delete_product_batch(batch_id: int, current_user=Depends(get_current_user)):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            """SELECT pb.*, p.branch_id FROM product_batches pb
               JOIN products p ON p.id = pb.product_id WHERE pb.id=%s FOR UPDATE""",
            (batch_id,),
        )
        batch = cur.fetchone()
        if not batch:
            raise HTTPException(status_code=404, detail="Batch not found")
        _assert_branch_access(current_user, batch["branch_id"])
        old_qty = int(batch["quantity"])
        set_batch_quantity(cur, batch_id, 0)
        if old_qty:
            log_movement(
                cur, batch["product_id"], batch["branch_id"], "adjustment",
                -old_qty, sync_product_from_batches(cur, batch["product_id"]),
                reference_type="batch", reason="Expiry lot removed",
                user_id=current_user.get("user_id"),
            )
        conn.commit()
        return {"ok": True}
    except HTTPException:
        conn.rollback()
        raise
    finally:
        conn.close()


# Max rows returned by list-style endpoints (raise if you outgrow this).
MAX_INVENTORY_ROWS = 100_000

# Unit value for stock valuation: prefer purchase cost; fall back to selling price when cost unset.
_STOCK_UNIT_VALUE = "COALESCE(NULLIF(cost, 0), price, 0)"
_STOCK_UNIT_VALUE_P = "COALESCE(NULLIF(p.cost, 0), p.price, 0)"
# products.stock is in sub-units when pack_size > 1; cost/price are per box (main unit).
_STOCK_QTY_BOXES_P = "(p.stock::float / GREATEST(COALESCE(NULLIF(p.pack_size, 0), 1), 1))"
# Branch-stock search results cap (keeps UI and DB fast)
BRANCH_STOCK_LIMIT = 1000


def _parse_search_terms(q: str) -> list[str]:
    if not q:
        return []
    normalized = q.replace("\n", ",").replace("\r", ",").replace(";", ",").replace("|", ",")
    return [t.strip() for t in normalized.split(",") if t.strip()]


def _branch_stock_term_clause(term: str) -> tuple[str, list]:
    """Match one term by exact/prefix barcode or by name contains."""
    from barcode_utils import barcode_lookup_candidates

    params: list = []
    name_like = f"%{term.strip()}%"
    parts = ["p.name_en ILIKE %s", "p.name_ar ILIKE %s"]
    params.extend([name_like, name_like])
    for cand in barcode_lookup_candidates(term):
        t_upper = cand.upper()
        prefix = f"{cand}%"
        parts.extend([
            "UPPER(COALESCE(p.barcode, '')) = %s",
            "UPPER(COALESCE(p.international_barcode, '')) = %s",
            "COALESCE(p.barcode, '') ILIKE %s",
            "COALESCE(p.international_barcode, '') ILIKE %s",
        ])
        params.extend([t_upper, t_upper, prefix, prefix])
    return f"({' OR '.join(parts)})", params


# ─── INVENTORY SUMMARY (accurate DB counts) ─────────────────────────────────

@router.get("/summary")
def inventory_summary(
    q: str = "",
    branch_id: Optional[int] = None,
    stock_filter: Optional[str] = None,
    category: Optional[str] = None,
    current_user=Depends(get_current_user),
    active_branch=Depends(get_active_branch_id),
):
    """Branch-scoped product counts for dashboard cards (not capped by list limits)."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if branch_id is not None and current_user.get("role") != "admin":
        if branch_id != current_user.get("branch_id"):
            raise HTTPException(status_code=403, detail="Cross-branch access denied")
    effective_branch = branch_id if branch_id is not None else active_branch
    where = ["p.active = true"]
    params: list = []
    if q:
        from barcode_utils import product_search_clause
        clause, clause_params = product_search_clause(q, table_prefix="p")
        where.append(clause)
        params += clause_params
    if effective_branch is not None:
        where.append("p.branch_id = %s")
        params.append(effective_branch)
    if category:
        where.append("p.category = %s")
        params.append(category)
    if stock_filter == "low":
        where.append("p.stock > 0 AND p.stock <= p.min_stock")
    elif stock_filter == "zero":
        where.append("p.stock <= 0")
    elif stock_filter == "ok":
        where.append("p.stock > p.min_stock")
    w = " AND ".join(where)
    # Stock value: box-equivalent qty × cost (or price); sub-unit stock divided by pack_size.
    cur.execute(
        f"""SELECT
              COUNT(*)::int AS total,
              COUNT(*) FILTER (WHERE p.stock <= 0)::int AS zero_stock,
              COUNT(*) FILTER (WHERE p.stock > 0 AND p.stock <= p.min_stock)::int AS low_stock,
              COALESCE(SUM({_STOCK_QTY_BOXES_P} * {_STOCK_UNIT_VALUE_P}), 0)::float AS stock_value
            FROM products p WHERE {w}""",
        params,
    )
    row = dict(cur.fetchone())
    conn.close()
    return row


# ─── LIST WITH FILTERS ─────────────────────────────────────────────────────

@router.get("/items")
def list_items(q: str = "", branch_id: Optional[int] = None,
               stock_filter: Optional[str] = None,
               category: Optional[str] = None,
               include_inactive: bool = False,
               load_all: bool = False,
               limit: int = 20000,
               current_user=Depends(get_current_user),
               active_branch=Depends(get_active_branch_id)):
    """stock_filter: 'low' | 'zero' | 'ok'. Requires search (q) or load_all=true."""
    if not load_all and not (q or "").strip():
        return []
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    where = []
    params = []
    if not include_inactive:
        where.append("p.active = true")
    if q:
        from barcode_utils import product_search_clause
        clause, clause_params = product_search_clause(q, table_prefix="p")
        where.append(clause)
        params += clause_params
    if branch_id is not None and current_user.get("role") != "admin":
        if branch_id != current_user.get("branch_id"):
            raise HTTPException(status_code=403, detail="Cross-branch access denied")
    effective_branch = branch_id if branch_id is not None else active_branch
    if effective_branch is not None:
        where.append("p.branch_id = %s")
        params.append(effective_branch)
    if category:
        where.append("p.category = %s")
        params.append(category)
    if stock_filter == "low":
        where.append("p.stock > 0 AND p.stock <= p.min_stock")
    elif stock_filter == "zero":
        where.append("p.stock <= 0")
    elif stock_filter == "ok":
        where.append("p.stock > p.min_stock")

    sql = (
        "SELECT p.*, b.name_en AS branch_name_en, b.name_ar AS branch_name_ar, "
        "COALESCE("
        "  (SELECT json_agg(json_build_object("
        "     'id', pb.id, 'expiry_date', pb.expiry_date, 'quantity', pb.quantity"
        "   ) ORDER BY COALESCE(pb.expiry_date, DATE '9999-12-31'), pb.id)"
        "   FROM product_batches pb WHERE pb.product_id = p.id AND pb.quantity > 0),"
        "  '[]'::json"
        ") AS batches "
        "FROM products p LEFT JOIN branches b ON p.branch_id = b.id"
    )
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY p.name_en LIMIT %s"
    params.append(max(1, min(limit, MAX_INVENTORY_ROWS)))
    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]


@router.get("/items/export")
def export_items(
    q: str = "",
    branch_id: Optional[int] = None,
    stock_filter: Optional[str] = None,
    category: Optional[str] = None,
    load_all: bool = False,
    current_user=Depends(get_current_user),
    active_branch=Depends(get_active_branch_id),
):
    rows = list_items(
        q=q,
        branch_id=branch_id,
        stock_filter=stock_filter,
        category=category,
        load_all=load_all,
        current_user=current_user,
        active_branch=active_branch,
    )
    headers = [
        "Barcode", "International Barcode", "Name EN", "Name AR", "Category", "Unit",
        "Price", "Cost", "Stock", "Min Stock", "Branch EN", "Branch AR", "Expiry",
    ]
    data = []
    for r in rows:
        exp = r.get("expiry_date")
        if not exp and r.get("batches"):
            batches = r["batches"] if isinstance(r["batches"], list) else []
            if batches:
                exp = batches[0].get("expiry_date")
        data.append([
            r.get("barcode"),
            r.get("international_barcode"),
            r.get("name_en"),
            r.get("name_ar"),
            r.get("category"),
            r.get("unit"),
            r.get("price"),
            r.get("cost"),
            r.get("stock"),
            r.get("min_stock"),
            r.get("branch_name_en"),
            r.get("branch_name_ar"),
            exp,
        ])
    return _xlsx_response(headers, data, "inventory_items.xlsx")


@router.get("/items/across-branches")
def search_across_branches(q: str,
                           current_user=Depends(get_current_user)):
    """Search by name/barcode across ALL branches — shows per-branch stock."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    s = f"%{q}%"
    cur.execute(
        """SELECT p.id, p.barcode, p.name_ar, p.name_en, p.unit, p.price,
                  p.stock, p.min_stock, p.branch_id,
                  b.name_en AS branch_name_en, b.name_ar AS branch_name_ar
           FROM products p LEFT JOIN branches b ON p.branch_id = b.id
           WHERE p.active = true
             AND (p.name_ar ILIKE %s OR p.name_en ILIKE %s OR p.barcode ILIKE %s)
           ORDER BY p.name_en, b.name_en LIMIT 200""",
        (s, s, s),
    )
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ─── CATEGORIES ─────────────────────────────────────────────────────────────

@router.get("/categories")
def list_categories(current_user=Depends(get_current_user)):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT DISTINCT category FROM products WHERE category IS NOT NULL AND category <> '' ORDER BY category"
    )
    rows = [r[0] for r in cur.fetchall()]
    conn.close()
    return rows


# ─── BRANCHES ───────────────────────────────────────────────────────────────

@router.get("/branches")
def list_branches(current_user=Depends(get_current_user)):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM branches ORDER BY name_en")
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ─── MANUAL STOCK ADJUSTMENT ────────────────────────────────────────────────

class AdjustmentRequest(BaseModel):
    product_id: int
    delta: Optional[int] = None  # positive or negative change
    set_to: Optional[int] = None  # exact stock in sub-units (pack-aware count)
    reason: str


@router.post("/adjustments")
def create_adjustment(req: AdjustmentRequest,
                      current_user=Depends(get_current_user)):
    if req.set_to is None and (req.delta is None or req.delta == 0):
        raise HTTPException(status_code=400, detail="Provide delta or set_to")
    if not req.reason or not req.reason.strip():
        raise HTTPException(status_code=400, detail="Reason is required")

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            "SELECT id, stock, branch_id, expiry_date FROM products WHERE id=%s FOR UPDATE",
            (req.product_id,),
        )
        product = cur.fetchone()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        _assert_branch_access(current_user, product["branch_id"])
        old_stock = int(product["stock"] or 0)
        if req.set_to is not None:
            if req.set_to < 0:
                raise HTTPException(status_code=400, detail="Stock cannot be negative")
            new_stock = set_product_stock_absolute(
                cur,
                req.product_id,
                product["branch_id"],
                int(req.set_to),
                product["expiry_date"],
            )
            delta = new_stock - old_stock
        else:
            delta = int(req.delta)
            if delta > 0:
                add_batch_stock(cur, req.product_id, product["branch_id"], delta, None)
            else:
                deduct_stock_fefo(cur, req.product_id, -delta, sellable_only=False)
            new_stock = sync_product_from_batches(cur, req.product_id)
        if new_stock < 0:
            raise HTTPException(status_code=400, detail="Adjustment would result in negative stock")
        if delta != 0:
            log_movement(
                cur, req.product_id, product["branch_id"], "adjustment",
                delta, new_stock,
                reference_type="adjustment", reason=req.reason,
                user_id=current_user.get("user_id"),
            )
        conn.commit()
        return {"ok": True, "new_stock": new_stock}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ─── STOCKTAKE (physical count reconciliation) ──────────────────────────

class StocktakeLot(BaseModel):
    expiry_date: Optional[date] = None
    quantity: int


class StocktakeLine(BaseModel):
    product_id: int
    counted: int
    expiry_date: Optional[date] = None
    category: Optional[str] = None
    # Optional per-expiry breakdown. When provided, it is the source of truth:
    # `counted` is derived from the sum of lot quantities and the product's
    # batches are rebuilt to match (mirrors how purchase orders record expiry).
    lots: Optional[List[StocktakeLot]] = None


class StocktakeRequest(BaseModel):
    items: List[StocktakeLine]
    note: Optional[str] = None


def _current_batches_signature(cur, product_id: int) -> list:
    """Sorted [(expiry_date_or_None, qty), ...] for the product's live batches."""
    cur.execute(
        """SELECT expiry_date, SUM(quantity)::int AS q
           FROM product_batches
           WHERE product_id = %s AND quantity > 0
           GROUP BY expiry_date""",
        (product_id,),
    )
    sig = [(r["expiry_date"], int(r["q"])) for r in cur.fetchall()]
    sig.sort(key=lambda t: (t[0] is None, t[0] or date(9999, 12, 31)))
    return sig


@router.post("/stocktake")
def apply_stocktake(req: StocktakeRequest,
                    current_user=Depends(get_current_user)):
    if not req.items:
        raise HTTPException(status_code=400, detail="No counted items provided")
    note = (req.note or "").strip()
    reason = f"Stocktake: {note}" if note else "Stocktake"
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # Resolve branch from first product
        cur.execute(
            "SELECT branch_id FROM products WHERE id=%s",
            (req.items[0].product_id,),
        )
        first = cur.fetchone()
        if not first or not first.get("branch_id"):
            raise HTTPException(status_code=400, detail="Invalid product")
        branch_id = first["branch_id"]
        _assert_branch_access(current_user, branch_id)

        cur.execute(
            """INSERT INTO stocktake_runs (branch_id, note, user_id)
               VALUES (%s,%s,%s) RETURNING id""",
            (branch_id, note or None, current_user.get("user_id")),
        )
        run_id = cur.fetchone()["id"]

        changes = []
        for line in req.items:
            cur.execute(
                """SELECT id, stock, branch_id, expiry_date, category, pack_size
                   FROM products WHERE id=%s FOR UPDATE""",
                (line.product_id,),
            )
            product = cur.fetchone()
            if not product:
                continue
            _assert_branch_access(current_user, product["branch_id"])
            old_stock = int(product["stock"])
            old_category = product.get("category")
            old_expiry = product.get("expiry_date")

            # Normalise per-expiry lots when provided (the multi-expiry path).
            use_lots = line.lots is not None
            norm_lots: List[StocktakeLot] = []
            if use_lots:
                for lot in line.lots or []:
                    if lot.quantity < 0:
                        raise HTTPException(status_code=400, detail="Lot quantity cannot be negative")
                    if lot.quantity > 0:
                        norm_lots.append(lot)
                counted = sum(int(lot.quantity) for lot in norm_lots)
                lot_expiries = [lot.expiry_date for lot in norm_lots if lot.expiry_date is not None]
                new_lot_expiry = min(lot_expiries) if lot_expiries else None
                new_sig = sorted(
                    [(lot.expiry_date, int(lot.quantity)) for lot in norm_lots],
                    key=lambda tpl: (tpl[0] is None, tpl[0] or date(9999, 12, 31)),
                )
                lots_change = new_sig != _current_batches_signature(cur, line.product_id)
            else:
                counted = line.counted
                if counted < 0:
                    raise HTTPException(status_code=400, detail="Counted quantity cannot be negative")
                lots_change = False
                new_lot_expiry = None

            delta = counted - old_stock
            if use_lots:
                expiry_change = lots_change
                new_expiry = new_lot_expiry
            else:
                expiry_change = line.expiry_date is not None and line.expiry_date != old_expiry
                new_expiry = line.expiry_date if expiry_change else old_expiry
            cat = (line.category or "").strip() if line.category is not None else None
            category_change = cat is not None and cat != (old_category or "")
            new_category = cat if category_change else old_category
            if delta == 0 and not expiry_change and not category_change:
                continue
            if category_change:
                cur.execute("UPDATE products SET category=%s WHERE id=%s", (cat, line.product_id))
            if use_lots:
                if delta != 0 or lots_change:
                    # Rebuild batches from scratch. Reset products.stock to 0 first so
                    # add_batch_stock's lazy migration cannot re-seed a phantom lot from
                    # the now-stale stock value.
                    cur.execute("DELETE FROM product_batches WHERE product_id=%s", (line.product_id,))
                    cur.execute("UPDATE products SET stock=0 WHERE id=%s", (line.product_id,))
                    for lot in norm_lots:
                        add_batch_stock(cur, line.product_id, product["branch_id"], int(lot.quantity), lot.expiry_date)
                    if not norm_lots:
                        sync_product_from_batches(cur, line.product_id)
            elif delta != 0 or expiry_change:
                cur.execute("DELETE FROM product_batches WHERE product_id=%s", (line.product_id,))
                cur.execute("UPDATE products SET stock=0 WHERE id=%s", (line.product_id,))
                exp = line.expiry_date if expiry_change else old_expiry
                qty = counted if delta != 0 else old_stock
                if qty > 0:
                    add_batch_stock(cur, line.product_id, product["branch_id"], qty, exp)
                else:
                    sync_product_from_batches(cur, line.product_id)
            if delta != 0:
                log_movement(
                    cur, line.product_id, product["branch_id"], "adjustment",
                    delta, counted,
                    reference_type="stocktake", reference_id=run_id,
                    reason=reason,
                    user_id=current_user.get("user_id"),
                )
            cur.execute(
                """INSERT INTO stocktake_lines
                   (run_id, product_id, old_stock, new_stock, delta,
                    old_category, new_category, old_expiry, new_expiry)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (
                    run_id, line.product_id, old_stock, counted, delta,
                    old_category, new_category, old_expiry, new_expiry,
                ),
            )
            changes.append({
                "product_id": line.product_id,
                "old_stock": old_stock,
                "new_stock": counted,
                "delta": delta,
            })
        if not changes:
            cur.execute("DELETE FROM stocktake_runs WHERE id=%s", (run_id,))
            conn.commit()
            return {"ok": True, "changed": 0, "run_id": None, "report": None}

        conn.commit()
        report = _build_stocktake_report(cur, run_id)
        return {"ok": True, "changed": len(changes), "run_id": run_id, "report": report}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/stocktake/runs")
def list_stocktake_runs(
    branch_id: Optional[int] = None,
    limit: int = Query(30, ge=1, le=200),
    current_user=Depends(get_current_user),
):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    where = []
    params: list = []
    if branch_id is not None:
        where.append("r.branch_id = %s")
        params.append(branch_id)
    elif current_user.get("role") != "admin":
        ub = current_user.get("branch_id")
        if ub is None:
            raise HTTPException(status_code=403, detail="No branch assigned")
        where.append("r.branch_id = %s")
        params.append(ub)
    sql = """
        SELECT r.id AS run_id, r.branch_id, r.note, r.created_at,
               b.name_en AS branch_name_en, b.name_ar AS branch_name_ar,
               u.name_en AS user_name_en, u.name_ar AS user_name_ar,
               COUNT(l.id)::int AS line_count,
               COUNT(l.id) FILTER (WHERE l.delta < 0)::int AS shortages_count,
               COUNT(l.id) FILTER (WHERE l.delta > 0)::int AS increases_count,
               COALESCE(SUM(l.delta) FILTER (WHERE l.delta < 0), 0)::int AS shortage_units,
               COALESCE(SUM(l.delta) FILTER (WHERE l.delta > 0), 0)::int AS increase_units
        FROM stocktake_runs r
        JOIN branches b ON b.id = r.branch_id
        LEFT JOIN users u ON u.id = r.user_id
        LEFT JOIN stocktake_lines l ON l.run_id = r.id
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " GROUP BY r.id, b.name_en, b.name_ar, u.name_en, u.name_ar"
    sql += " ORDER BY r.created_at DESC LIMIT %s"
    params.append(limit)
    cur.execute(sql, params)
    rows = []
    for r in cur.fetchall():
        row = dict(r)
        created = row.get("created_at")
        if created and hasattr(created, "isoformat"):
            row["created_at"] = created.isoformat()
        rows.append(row)
    conn.close()
    return rows


@router.get("/stocktake/runs/{run_id}")
def get_stocktake_report(run_id: int, current_user=Depends(get_current_user)):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT branch_id FROM stocktake_runs WHERE id=%s", (run_id,))
        run = cur.fetchone()
        if not run:
            raise HTTPException(status_code=404, detail="Stocktake run not found")
        _assert_branch_access(current_user, run["branch_id"])
        return _build_stocktake_report(cur, run_id)
    finally:
        conn.close()


def _stocktake_line_export_row(ln: dict) -> list:
    pack = ln.get("pack_size") or 1
    return [
        ln.get("name_en"), ln.get("name_ar"), ln.get("barcode"), ln.get("category"),
        ln.get("old_stock"), ln.get("new_stock"), ln.get("delta"),
        ln.get("variance_major"), ln.get("variance_sub_fraction"),
        ln.get("unit"), ln.get("sub_unit"),
        ln.get("old_expiry"), ln.get("new_expiry"),
        ln.get("old_category"), ln.get("new_category"),
    ]


STOCKTAKE_LINE_HEADERS = [
    "Product EN", "Product AR", "Barcode", "Category",
    "System qty", "Counted qty", "Delta (sub-units)",
    "Var. boxes", "Var. strip fraction",
    "Unit", "Sub-unit", "Old expiry", "New expiry",
    "Old category", "New category",
]


@router.get("/stocktake/runs/{run_id}/export")
def export_stocktake_report(run_id: int, current_user=Depends(get_current_user)):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT branch_id FROM stocktake_runs WHERE id=%s", (run_id,))
        run = cur.fetchone()
        if not run:
            raise HTTPException(status_code=404, detail="Stocktake run not found")
        _assert_branch_access(current_user, run["branch_id"])
        report = _build_stocktake_report(cur, run_id)
    finally:
        conn.close()

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    sheets = [
        ("Shortages", report.get("shortages", [])),
        ("Increases", report.get("increases", [])),
        ("Other changes", report.get("other_changes", [])),
    ]
    for title, rows in sheets:
        ws = wb.create_sheet(title[:31])
        ws.append(STOCKTAKE_LINE_HEADERS)
        for ln in rows:
            ws.append([_xlsx_safe(v) for v in _stocktake_line_export_row(ln)])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"stocktake_report_{run_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── CLEAR BRANCH HISTORY (admin, password-protected) ────────────────────

class ClearHistoryRequest(BaseModel):
    branch_id: int
    password: str


@router.post("/clear-branch-history")
def clear_branch_history(req: ClearHistoryRequest,
                          current_user=Depends(get_current_user)):
    """Wipe a branch's sales history, stock transfers, and movements; reset
    product stock to zero. Admin only; requires the caller's password as guard."""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    from auth import verify_password
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT password_hash FROM users WHERE id=%s",
                    (current_user.get("user_id"),))
        u = cur.fetchone()
        if not u or not verify_password(req.password, u["password_hash"]):
            raise HTTPException(status_code=403, detail="Incorrect password")

        cur.execute("SELECT id, name_en FROM branches WHERE id=%s", (req.branch_id,))
        b = cur.fetchone()
        if not b:
            raise HTTPException(status_code=404, detail="Branch not found")

        cur.execute("DELETE FROM stock_movements WHERE branch_id=%s", (req.branch_id,))
        moves = cur.rowcount
        cur.execute(
            """DELETE FROM stock_transfers
               WHERE from_branch_id=%s OR to_branch_id=%s""",
            (req.branch_id, req.branch_id),
        )
        transfers = cur.rowcount
        cur.execute("DELETE FROM returns WHERE branch_id=%s", (req.branch_id,))
        rets = cur.rowcount
        cur.execute(
            """DELETE FROM customer_payments
               WHERE invoice_id IN (SELECT id FROM invoices WHERE branch_id=%s)""",
            (req.branch_id,),
        )
        pays = cur.rowcount
        cur.execute("DELETE FROM invoices WHERE branch_id=%s", (req.branch_id,))
        invs = cur.rowcount
        cur.execute("UPDATE products SET stock=0 WHERE branch_id=%s", (req.branch_id,))
        prods = cur.rowcount
        conn.commit()
        return {"ok": True, "branch_id": req.branch_id,
                "deleted_movements": moves, "deleted_transfers": transfers,
                "deleted_returns": rets, "deleted_payments": pays,
                "deleted_invoices": invs, "reset_products": prods}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ─── STOCK MOVEMENT REPORT ─────────────────────────────────────────────────

@router.get("/movements")
def list_movements(product_id: Optional[int] = None,
                   branch_id: Optional[int] = None,
                   movement_type: Optional[str] = None,
                   start_date: Optional[date] = None,
                   end_date: Optional[date] = None,
                   limit: int = 200,
                   current_user=Depends(get_current_user)):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    where = []
    params = []
    if product_id:
        where.append("m.product_id = %s"); params.append(product_id)
    if branch_id:
        where.append("m.branch_id = %s"); params.append(branch_id)
    if movement_type:
        where.append("m.movement_type = %s"); params.append(movement_type)
    if start_date:
        where.append("DATE(m.created_at) >= %s"); params.append(start_date)
    if end_date:
        where.append("DATE(m.created_at) <= %s"); params.append(end_date)

    sql = (
        """SELECT m.*, p.name_ar AS product_name_ar, p.name_en AS product_name_en,
                  p.barcode, u.name_en AS user_name_en, u.name_ar AS user_name_ar,
                  b.name_en AS branch_name_en, b.name_ar AS branch_name_ar
           FROM stock_movements m
           LEFT JOIN products p ON m.product_id = p.id
           LEFT JOIN users u ON m.user_id = u.id
           LEFT JOIN branches b ON m.branch_id = b.id"""
    )
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY m.created_at DESC LIMIT %s"
    params.append(limit)
    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]


@router.get("/movements/export")
def export_movements(
    product_id: Optional[int] = None,
    branch_id: Optional[int] = None,
    movement_type: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    limit: int = 5000,
    current_user=Depends(get_current_user),
):
    rows = list_movements(
        product_id=product_id,
        branch_id=branch_id,
        movement_type=movement_type,
        start_date=start_date,
        end_date=end_date,
        limit=limit,
        current_user=current_user,
    )
    headers = [
        "Date", "Product EN", "Product AR", "Type", "Qty", "Balance After",
        "Reason", "User EN", "User AR", "Branch EN",
    ]
    data = [
        [
            r.get("created_at"),
            r.get("product_name_en"),
            r.get("product_name_ar"),
            r.get("movement_type"),
            r.get("quantity"),
            r.get("balance_after"),
            r.get("reason"),
            r.get("user_name_en"),
            r.get("user_name_ar"),
            r.get("branch_name_en"),
        ]
        for r in rows
    ]
    return _xlsx_response(headers, data, "stock_movements.xlsx")


# ─── FAST / SLOW / DEAD CLASSIFICATION ─────────────────────────────────────

@router.get("/velocity")
def velocity_classification(days: int = 90,
                            date_from: Optional[str] = None,
                            date_to: Optional[str] = None,
                            current_user=Depends(get_current_user)):
    """
    Classify items by sales velocity over a period.
    Uses a custom [date_from, date_to] range when both are supplied,
    otherwise the last N days.
    - fast:  >= 10 units sold
    - slow:  1-9 units sold
    - dead:  0 units sold
    """
    if date_from or date_to:
        if not (date_from and date_to):
            raise HTTPException(
                status_code=400,
                detail="Both date_from and date_to are required for a custom range",
            )
        try:
            d_from = date.fromisoformat(date_from)
            d_to = date.fromisoformat(date_to)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format, expected YYYY-MM-DD")
        if d_from > d_to:
            raise HTTPException(status_code=400, detail="date_from must be on or before date_to")
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if date_from and date_to:
        sold_filter = ("WHERE i.created_at >= %s::date "
                       "AND i.created_at < (%s::date + INTERVAL '1 day')")
        filter_params = [date_from, date_to]
    else:
        sold_filter = "WHERE i.created_at >= NOW() - (%s * INTERVAL '1 day')"
        filter_params = [days]
    cur.execute(
        f"""SELECT p.id, p.name_ar, p.name_en, p.barcode, p.stock, p.unit, p.price,
                  COALESCE(SUM(ii.quantity) FILTER (
                      {sold_filter}
                  ), 0) AS sold_qty
           FROM products p
           LEFT JOIN invoice_items ii ON ii.product_id = p.id
           LEFT JOIN invoices i ON ii.invoice_id = i.id AND i.status='completed'
           WHERE p.active = true
           GROUP BY p.id
           ORDER BY sold_qty DESC""",
        filter_params,
    )
    rows = cur.fetchall()
    conn.close()
    result = []
    for r in rows:
        sold = int(r["sold_qty"] or 0)
        if sold >= 10:
            cls = "fast"
        elif sold >= 1:
            cls = "slow"
        else:
            cls = "dead"
        item = dict(r)
        item["sold_qty"] = sold
        item["classification"] = cls
        result.append(item)
    return result


@router.get("/velocity/export")
def export_velocity(
    days: int = 90,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user=Depends(get_current_user),
):
    rows = velocity_classification(
        days=days, date_from=date_from, date_to=date_to, current_user=current_user,
    )
    headers = ["Name EN", "Name AR", "Barcode", "Stock", "Sold Qty", "Classification", "Unit", "Price"]
    data = [
        [
            r.get("name_en"), r.get("name_ar"), r.get("barcode"), r.get("stock"),
            r.get("sold_qty"), r.get("classification"), r.get("unit"), r.get("price"),
        ]
        for r in rows
    ]
    return _xlsx_response(headers, data, "inventory_velocity.xlsx")


# ─── CONSUMPTION-BASED MIN STOCK SUGGESTION ────────────────────────────────

DEFAULT_SUB_UNIT = "piece"


def _row_get(r, *keys):
    """Return the first non-empty value among the given (lowercased) header keys."""
    for k in keys:
        if k in r and r[k] not in (None, ""):
            return r[k]
    return None


@router.get("/bulk-template")
def bulk_template(current_user=Depends(get_current_user)):
    """Download a blank Excel template for bulk item upload."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"
    headers = ["Code", "International Barcode", "Material Name", "Name (Arabic)", "Unit", "Small Unit",
               "Small Unit Quantity Per Unit", "Quantity",
               "Sales Price", "Cost", "Category", "Min Stock", "Expiry Date"]
    ws.append(headers)
    ws.append(["1234567890123", "5000112637922", "Panadol Extra 48 Tab", "بانادول اكسترا 48 قرص",
               "Box", "Strip", 4, 100,
               116.00, 80.00, "Painkillers", 10, "2027-12-31"])
    ws.append(["7654321098765", "8901234567890", "Augmentin 1g", "اوجمنتين 1 جم",
               "Box", "Tablet", 14, 50,
               180.00, 130.00, "Antibiotics", 5, "2026-06-30"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="items_template.xlsx"'},
    )


@router.post("/bulk-upload")
async def bulk_upload(file: UploadFile = File(...),
                      current_user=Depends(get_current_user),
                      active_branch=Depends(get_active_branch_id)):
    """Bulk import items from Excel/CSV. Existing barcodes are updated."""
    from openpyxl import load_workbook
    content = await file.read()
    name = (file.filename or "").lower()

    rows = []
    if name.endswith(".csv"):
        import csv
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    else:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(r):
                continue
            row = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
            rows.append(row)

    rows = [{(str(k).strip().lower() if k is not None else ""): v
             for k, v in row.items()} for row in rows]

    inserted = updated = errors = 0
    error_details = []
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    branch_id = active_branch if active_branch is not None else current_user.get("branch_id")
    if branch_id is None:
        conn.close()
        raise HTTPException(status_code=400, detail="Select a specific branch before uploading items")
    user_id = current_user.get("user_id")

    for idx, r in enumerate(rows, start=2):
        cur.execute("SAVEPOINT row_sp")
        try:
            name_en = str(_row_get(r, "name_en", "name", "material name", "item name") or "").strip()
            if not name_en:
                raise ValueError("Material Name is required")
            name_ar = str(_row_get(r, "name_ar", "arabic name", "name (arabic)") or "").strip() or name_en
            barcode = str(_row_get(r, "barcode", "code", "material code", "item code") or "").strip() or None
            intl_barcode = str(_row_get(r, "international barcode", "international_barcode",
                                        "intl barcode", "global barcode", "gtin", "ean") or "").strip() or None
            category = str(_row_get(r, "category") or "").strip() or None
            unit = str(_row_get(r, "unit", "material unit", "big unit", "main unit") or "box").strip()
            price = float(_row_get(r, "price", "sales price", "selling price", "material price") or 0)
            cost_val = _row_get(r, "cost", "cost price", "purchase price")
            cost = float(cost_val) if cost_val not in (None, "") else None
            if cost is None and price > 0:
                cost = price
            min_stock = int(float(_row_get(r, "min_stock", "min stock", "minimum stock") or 5))

            pack_raw = _row_get(r, "pack_size", "pack size",
                                "small unit quantity per unit", "small unit qty per unit",
                                "small unit per unit", "small unit quantity",
                                "units per unit", "quantity per unit", "qty per unit",
                                "number of small unit", "number of small units",
                                "small units", "units per pack", "units per box",
                                "units per 1 box", "units per big unit")
            sub_unit_name = str(_row_get(r, "small unit", "small_unit", "sub_unit",
                                         "sub unit", "small unit name",
                                         "unit classification", "small unit type") or "").strip()

            if pack_raw in (None, "") and sub_unit_name:
                try:
                    pack_raw = float(sub_unit_name)
                    sub_unit_name = ""
                except ValueError:
                    pass

            pack_size = int(float(pack_raw)) if pack_raw not in (None, "") else 1
            if pack_size < 1:
                pack_size = 1
            sub_unit_name = sub_unit_name.lower()

            qty_raw = _row_get(r, "stock", "quantity", "qty")
            qty_provided = qty_raw not in (None, "")
            qty_big = float(qty_raw) if qty_provided else 0.0

            from datetime import datetime as _dt, date as _date
            expiry_date = None
            exp_raw = _row_get(r, "expiry", "expiry date", "expiry_date",
                               "expiration", "expiration date", "exp date", "exp")
            if exp_raw not in (None, ""):
                if isinstance(exp_raw, _dt):
                    expiry_date = exp_raw.date()
                elif isinstance(exp_raw, _date):
                    expiry_date = exp_raw
                else:
                    s = str(exp_raw).strip()
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                        try:
                            expiry_date = _dt.strptime(s, fmt).date()
                            break
                        except ValueError:
                            continue

            if pack_size > 1:
                sub_unit = sub_unit_name or DEFAULT_SUB_UNIT
                sub_price = round(price / pack_size, 2) if price else None
                stock = int(round(qty_big * pack_size))
            else:
                sub_unit = None
                sub_price = None
                stock = int(round(qty_big))

            existing = None
            if barcode:
                if branch_id is None:
                    cur.execute("SELECT id, stock FROM products WHERE barcode=%s AND branch_id IS NULL", (barcode,))
                else:
                    cur.execute("SELECT id, stock FROM products WHERE barcode=%s AND branch_id=%s", (barcode, branch_id))
                existing = cur.fetchone()

            if existing:
                cur.execute(
                    """UPDATE products SET name_en=%s, name_ar=%s, category=%s, unit=%s,
                       price=%s, cost=%s, min_stock=%s, pack_size=%s, sub_unit=%s,
                       sub_price=%s, active=true,
                       international_barcode=COALESCE(%s, international_barcode)
                       WHERE id=%s""",
                    (name_en, name_ar, category, unit, price, cost, min_stock,
                     pack_size, sub_unit, sub_price, intl_barcode, existing["id"]),
                )
                # Stock & expiry live in the FEFO batch ledger. Only touch them when
                # the row actually carries a quantity, so a metadata-only re-import
                # never silently zeroes stock or wipes expiry batches.
                if qty_provided:
                    old_q = int(existing["stock"])
                    delta = stock - old_q
                    if delta > 0:
                        add_batch_stock(cur, existing["id"], branch_id, delta, expiry_date)
                    elif delta < 0:
                        deduct_stock_fefo(cur, existing["id"], -delta, sellable_only=False)
                    elif expiry_date is not None:
                        # Quantity unchanged but a single expiry was supplied:
                        # rebuild the lot so the expiry lands in the batch ledger.
                        cur.execute("DELETE FROM product_batches WHERE product_id=%s", (existing["id"],))
                        if stock > 0:
                            add_batch_stock(cur, existing["id"], branch_id, stock, expiry_date)
                        else:
                            sync_product_from_batches(cur, existing["id"])
                    if delta != 0:
                        log_movement(
                            cur, existing["id"], branch_id, "adjustment",
                            delta, stock,
                            reference_type="bulk_upload", reason="Bulk upload sync",
                            user_id=user_id,
                        )
                elif expiry_date is not None:
                    cur.execute(
                        "UPDATE products SET expiry_date=COALESCE(%s, expiry_date) WHERE id=%s",
                        (expiry_date, existing["id"]),
                    )
                updated += 1
            else:
                cur.execute(
                    """INSERT INTO products (barcode, international_barcode, name_ar, name_en, category, unit,
                       price, cost, stock, min_stock, pack_size, sub_unit, sub_price, expiry_date, branch_id)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                    (barcode, intl_barcode, name_ar, name_en, category, unit,
                     price, cost, 0, min_stock, pack_size, sub_unit, sub_price, expiry_date, branch_id),
                )
                new_id = cur.fetchone()["id"]
                # Seed the opening quantity through the batch ledger so the expiry
                # is recorded as a FEFO lot (consistent with PO receive & stocktake).
                if qty_provided and stock > 0:
                    add_batch_stock(cur, new_id, branch_id, stock, expiry_date)
                    log_movement(
                        cur, new_id, branch_id, "initial",
                        stock, stock,
                        reference_type="bulk_upload", reason="Bulk upload",
                        user_id=user_id,
                    )
                inserted += 1
            cur.execute("RELEASE SAVEPOINT row_sp")
        except Exception as e:
            cur.execute("ROLLBACK TO SAVEPOINT row_sp")
            errors += 1
            error_details.append(f"Row {idx}: {e}")

    conn.commit()
    conn.close()
    return {
        "inserted": inserted, "updated": updated, "errors": errors,
        "error_details": error_details[:50],
    }


@router.get("/consumption-alerts")
def consumption_alerts(days: int = 30,
                       coverage_days: int = 7,
                       current_user=Depends(get_current_user)):
    """
    Alert when current stock < (avg daily consumption × coverage_days).
    Based on last `days` days of sales.
    """
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """SELECT p.id, p.name_ar, p.name_en, p.barcode, p.stock, p.min_stock,
                  COALESCE(s.sold, 0)::float / %s AS avg_daily
           FROM products p
           LEFT JOIN (
               SELECT ii.product_id, SUM(ii.quantity) AS sold
               FROM invoice_items ii
               JOIN invoices i ON ii.invoice_id = i.id
               WHERE i.status='completed'
                 AND i.created_at >= NOW() - (%s || ' days')::interval
               GROUP BY ii.product_id
           ) s ON s.product_id = p.id
           WHERE p.active = true""",
        (days, days),
    )
    rows = cur.fetchall()
    conn.close()
    alerts = []
    for r in rows:
        avg = float(r["avg_daily"] or 0)
        suggested_min = round(avg * coverage_days)
        if avg > 0 and int(r["stock"]) < suggested_min:
            d = dict(r)
            d["avg_daily"] = round(avg, 2)
            d["suggested_min"] = suggested_min
            d["days_remaining"] = round(int(r["stock"]) / avg, 1) if avg > 0 else None
            alerts.append(d)
    return alerts


@router.get("/consumption-alerts/export")
def export_consumption_alerts(
    days: int = 30,
    coverage_days: int = 7,
    current_user=Depends(get_current_user),
):
    rows = consumption_alerts(days=days, coverage_days=coverage_days, current_user=current_user)
    headers = ["Name EN", "Name AR", "Barcode", "Stock", "Avg Daily", "Days Left", "Suggested Min"]
    data = [
        [
            r.get("name_en"), r.get("name_ar"), r.get("barcode"), r.get("stock"),
            r.get("avg_daily"), r.get("days_remaining"), r.get("suggested_min"),
        ]
        for r in rows
    ]
    return _xlsx_response(headers, data, "inventory_alerts.xlsx")


# ─── SMART STOCK REALLOCATION ───────────────────────────────────────────────

@router.get("/reallocation-suggestions", dependencies=[Depends(requires_feature("stock_reallocation"))])
def reallocation_suggestions(
    days: int = Query(30, ge=7, le=365),
    coverage_days: int = Query(7, ge=1, le=60),
    surplus_factor: float = Query(2.0, ge=1.2, le=5.0),
    current_user=Depends(get_current_user),
):
    """
    Suggest inter-branch transfers based on consumption vs stock.
    Matches products by barcode across branches.
    """
    if current_user.get("role") not in ("admin", "pharmacist"):
        raise HTTPException(403, "Admin or pharmacist required")

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            """
            WITH branch_sales AS (
              SELECT p.barcode,
                     p.branch_id,
                     p.id AS product_id,
                     p.name_en,
                     p.name_ar,
                     COALESCE(p.stock, 0)::float AS stock,
                     COALESCE(SUM(ii.quantity), 0)::float AS sold_qty
              FROM products p
              LEFT JOIN invoice_items ii ON ii.product_id = p.id
              LEFT JOIN invoices i ON i.id = ii.invoice_id
                AND i.status = 'completed'
                AND i.created_at >= NOW() - (%s * INTERVAL '1 day')
              WHERE p.active = true
                AND p.barcode IS NOT NULL
                AND TRIM(p.barcode) <> ''
              GROUP BY p.id, p.barcode, p.branch_id, p.name_en, p.name_ar, p.stock
            ),
            metrics AS (
              SELECT bs.*,
                     b.name_en AS branch_name_en,
                     b.name_ar AS branch_name_ar,
                     ROUND((bs.sold_qty / %s)::numeric, 2)::float AS avg_daily,
                     CASE WHEN bs.sold_qty > 0
                       THEN ROUND((bs.stock / (bs.sold_qty / %s))::numeric, 1)
                       ELSE NULL END AS days_cover
              FROM branch_sales bs
              JOIN branches b ON b.id = bs.branch_id
            )
            SELECT * FROM metrics
            ORDER BY barcode, branch_id
            """,
            [days, days, days],
        )
        rows = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()

    by_barcode: dict[str, list] = {}
    for r in rows:
        bc = (r.get("barcode") or "").strip()
        if not bc:
            continue
        by_barcode.setdefault(bc, []).append(r)

    suggestions: list[dict] = []
    target_stock = coverage_days  # multiplier applied via avg_daily

    for barcode, branches in by_barcode.items():
        if len(branches) < 2:
            continue
        needs = []
        surpluses = []
        for b in branches:
            avg = float(b.get("avg_daily") or 0)
            stock = float(b.get("stock") or 0)
            if avg <= 0:
                continue
            need_qty = max(0, round(avg * target_stock - stock))
            surplus_qty = max(0, round(stock - avg * target_stock * surplus_factor))
            if need_qty >= 1:
                needs.append({**b, "need_qty": int(need_qty)})
            if surplus_qty >= 1:
                surpluses.append({**b, "surplus_qty": int(surplus_qty)})

        needs.sort(key=lambda x: -x["need_qty"])
        surpluses.sort(key=lambda x: -x["surplus_qty"])

        for need in needs:
            remaining = need["need_qty"]
            for surplus in surpluses:
                if surplus["branch_id"] == need["branch_id"]:
                    continue
                if surplus["surplus_qty"] <= 0:
                    continue
                qty = min(remaining, surplus["surplus_qty"])
                if qty < 1:
                    continue
                suggestions.append({
                    "barcode": barcode,
                    "name_en": need.get("name_en") or surplus.get("name_en"),
                    "name_ar": need.get("name_ar") or surplus.get("name_ar"),
                    "from_branch_id": surplus["branch_id"],
                    "from_branch_name_en": surplus["branch_name_en"],
                    "from_branch_name_ar": surplus["branch_name_ar"],
                    "from_product_id": surplus["product_id"],
                    "from_stock": surplus["stock"],
                    "to_branch_id": need["branch_id"],
                    "to_branch_name_en": need["branch_name_en"],
                    "to_branch_name_ar": need["branch_name_ar"],
                    "to_product_id": need["product_id"],
                    "to_stock": need["stock"],
                    "suggested_qty": int(qty),
                    "need_branch_avg_daily": need.get("avg_daily"),
                    "from_branch_avg_daily": surplus.get("avg_daily"),
                    "priority_score": round(need["need_qty"] * float(need.get("avg_daily") or 1), 2),
                })
                surplus["surplus_qty"] -= qty
                remaining -= qty
                if remaining <= 0:
                    break

    suggestions.sort(key=lambda x: -x["priority_score"])
    return {
        "days": days,
        "coverage_days": coverage_days,
        "surplus_factor": surplus_factor,
        "count": len(suggestions),
        "suggestions": suggestions,
    }


@router.get("/reallocation-suggestions/export", dependencies=[Depends(requires_feature("stock_reallocation"))])
def export_reallocation_suggestions(
    days: int = 30,
    coverage_days: int = 7,
    surplus_factor: float = 2.0,
    current_user=Depends(get_current_user),
):
    report = reallocation_suggestions(
        days=days, coverage_days=coverage_days, surplus_factor=surplus_factor,
        current_user=current_user,
    )
    headers = [
        "Barcode", "Name EN", "From branch", "From stock", "To branch", "To stock",
        "Suggested qty", "Need avg/day", "From avg/day", "Priority",
    ]
    data = [
        [
            s["barcode"], s.get("name_en"),
            s.get("from_branch_name_en"), s.get("from_stock"),
            s.get("to_branch_name_en"), s.get("to_stock"),
            s.get("suggested_qty"), s.get("need_branch_avg_daily"),
            s.get("from_branch_avg_daily"), s.get("priority_score"),
        ]
        for s in report["suggestions"]
    ]
    return _xlsx_response(headers, data, "stock_reallocation.xlsx")


# ─── STOCK TRANSFERS ────────────────────────────────────────────────────────

class TransferItemIn(BaseModel):
    product_id: int
    quantity: int


class TransferRequest(BaseModel):
    from_branch_id: int
    to_branch_id: int
    items: List[TransferItemIn]
    notes: Optional[str] = None


def _assert_can_transfer_from(user, from_branch_id: int):
    if user.get("role") == "admin":
        return
    if user.get("branch_id") != from_branch_id:
        raise HTTPException(status_code=403, detail="You can only transfer from your own branch")


@router.post("/transfers")
def create_transfer(req: TransferRequest, current_user=Depends(get_current_user)):
    if req.from_branch_id == req.to_branch_id:
        raise HTTPException(status_code=400, detail="Source and destination must differ")
    if not req.items:
        raise HTTPException(status_code=400, detail="At least one item is required")
    _assert_can_transfer_from(current_user, req.from_branch_id)

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # validate branches
        cur.execute("SELECT id FROM branches WHERE id IN (%s, %s)",
                    (req.from_branch_id, req.to_branch_id))
        if len(cur.fetchall()) != 2:
            raise HTTPException(status_code=400, detail="Invalid branch(es)")

        cur.execute("SELECT nextval('stock_transfer_seq') AS n")
        seq_n = cur.fetchone()["n"]
        transfer_number = f"TRN-{date.today().strftime('%Y%m%d')}-{int(seq_n):04d}"

        cur.execute(
            """INSERT INTO stock_transfers
               (transfer_number, from_branch_id, to_branch_id, status, notes, created_by)
               VALUES (%s,%s,%s,'in_transit',%s,%s) RETURNING id""",
            (transfer_number, req.from_branch_id, req.to_branch_id,
             req.notes, current_user.get("user_id")),
        )
        transfer_id = cur.fetchone()["id"]

        for it in req.items:
            if it.quantity <= 0:
                raise HTTPException(status_code=400, detail="Quantity must be positive")
            cur.execute(
                "SELECT id, stock, branch_id, barcode, name_ar, name_en, unit, sub_unit, pack_size FROM products WHERE id=%s FOR UPDATE",
                (it.product_id,),
            )
            p = cur.fetchone()
            if not p:
                raise HTTPException(status_code=404, detail=f"Product {it.product_id} not found")
            if p["branch_id"] != req.from_branch_id:
                raise HTTPException(
                    status_code=400,
                    detail=f"Product {p['name_en']} is not in source branch {req.from_branch_id}",
                )
            if int(p["stock"]) < it.quantity:
                raise HTTPException(
                    status_code=400,
                    detail=f"Insufficient stock for {p['name_en']} (have {p['stock']}, need {it.quantity})",
                )
            deduct_stock_fefo(cur, p["id"], it.quantity, sellable_only=False)
            new_stock = sync_product_from_batches(cur, p["id"])
            # Use sub_unit when pack_size > 1 (stock is tracked in sub-units), else main unit.
            unit_label = (p.get("sub_unit") if (p.get("pack_size") or 1) > 1 and p.get("sub_unit")
                          else p.get("unit") or "unit")
            cur.execute(
                """INSERT INTO stock_transfer_items
                   (transfer_id, source_product_id, barcode, product_name_ar, product_name_en, quantity, unit_label)
                   VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                (transfer_id, p["id"], p["barcode"], p["name_ar"], p["name_en"], it.quantity, unit_label),
            )
            log_movement(
                cur, p["id"], req.from_branch_id, "transfer_out",
                -it.quantity, new_stock,
                reference_type="transfer", reference_id=transfer_id,
                reason=f"Transfer {transfer_number} → branch {req.to_branch_id}",
                user_id=current_user.get("user_id"),
            )
        conn.commit()
        return {"ok": True, "transfer_id": transfer_id, "transfer_number": transfer_number}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/transfers")
def list_transfers(status: Optional[str] = None,
                   current_user=Depends(get_current_user)):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    where = []
    params = []
    if status:
        where.append("t.status = %s")
        params.append(status)
    if current_user.get("role") != "admin":
        ub = current_user.get("branch_id")
        where.append("(t.from_branch_id = %s OR t.to_branch_id = %s)")
        params += [ub, ub]
    sql = """SELECT t.*,
                    bf.name_en AS from_name_en, bf.name_ar AS from_name_ar,
                    bt.name_en AS to_name_en, bt.name_ar AS to_name_ar,
                    u.name_en AS created_by_name_en, u.name_ar AS created_by_name_ar
             FROM stock_transfers t
             LEFT JOIN branches bf ON t.from_branch_id = bf.id
             LEFT JOIN branches bt ON t.to_branch_id = bt.id
             LEFT JOIN users u ON t.created_by = u.id"""
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY t.created_at DESC LIMIT 200"
    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]


@router.get("/transfers/{transfer_id}")
def get_transfer(transfer_id: int, current_user=Depends(get_current_user)):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """SELECT t.*,
                  bf.name_en AS from_name_en, bf.name_ar AS from_name_ar,
                  bt.name_en AS to_name_en, bt.name_ar AS to_name_ar
           FROM stock_transfers t
           LEFT JOIN branches bf ON t.from_branch_id = bf.id
           LEFT JOIN branches bt ON t.to_branch_id = bt.id
           WHERE t.id=%s""",
        (transfer_id,),
    )
    t = cur.fetchone()
    if not t:
        conn.close()
        raise HTTPException(status_code=404, detail="Transfer not found")
    if current_user.get("role") != "admin":
        ub = current_user.get("branch_id")
        if ub not in (t["from_branch_id"], t["to_branch_id"]):
            conn.close()
            raise HTTPException(status_code=403, detail="Not accessible")
    cur.execute("SELECT * FROM stock_transfer_items WHERE transfer_id=%s ORDER BY id", (transfer_id,))
    items = cur.fetchall()
    conn.close()
    out = dict(t)
    out["items"] = [dict(i) for i in items]
    return out


@router.post("/transfers/{transfer_id}/receive")
def receive_transfer(transfer_id: int, current_user=Depends(get_current_user)):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT * FROM stock_transfers WHERE id=%s FOR UPDATE", (transfer_id,))
        t = cur.fetchone()
        if not t:
            raise HTTPException(status_code=404, detail="Transfer not found")
        if t["status"] != "in_transit":
            raise HTTPException(status_code=400, detail=f"Cannot receive transfer in status '{t['status']}'")
        # Authorization: admin OR member of destination branch
        if current_user.get("role") != "admin":
            if current_user.get("branch_id") != t["to_branch_id"]:
                raise HTTPException(status_code=403, detail="Only destination branch can receive")
        to_branch = t["to_branch_id"]

        cur.execute("SELECT * FROM stock_transfer_items WHERE transfer_id=%s", (transfer_id,))
        items = cur.fetchall()
        for it in items:
            cur.execute("SELECT * FROM products WHERE id=%s", (it["source_product_id"],))
            src = cur.fetchone()
            if not src:
                raise HTTPException(status_code=404, detail="Source product not found")
            dest = None
            if it["barcode"]:
                cur.execute(
                    "SELECT id, stock FROM products WHERE barcode=%s AND branch_id=%s FOR UPDATE",
                    (it["barcode"], to_branch),
                )
                dest = cur.fetchone()
            if not dest:
                cur.execute(
                    """INSERT INTO products
                       (barcode, name_ar, name_en, category, unit, price, cost,
                        stock, min_stock, expiry_date, branch_id, active)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,0,%s,%s,%s,true)
                       ON CONFLICT (barcode, branch_id) DO NOTHING
                       RETURNING id, stock""",
                    (src["barcode"], src["name_ar"], src["name_en"], src["category"],
                     src["unit"], src["price"], src["cost"],
                     src["min_stock"], src["expiry_date"], to_branch),
                )
                dest = cur.fetchone()
                if not dest:
                    # A concurrent receive inserted it first — re-fetch with lock.
                    cur.execute(
                        "SELECT id, stock FROM products WHERE barcode=%s AND branch_id=%s FOR UPDATE",
                        (src["barcode"], to_branch),
                    )
                    dest = cur.fetchone()
                    if not dest:
                        raise HTTPException(
                            status_code=409,
                            detail=f"Could not resolve destination product for barcode {src['barcode']}",
                        )
            add_batch_stock(cur, dest["id"], to_branch, int(it["quantity"]), src.get("expiry_date"))
            new_stock = sync_product_from_batches(cur, dest["id"])
            cur.execute(
                "UPDATE stock_transfer_items SET dest_product_id=%s WHERE id=%s",
                (dest["id"], it["id"]),
            )
            log_movement(
                cur, dest["id"], to_branch, "transfer_in",
                int(it["quantity"]), new_stock,
                reference_type="transfer", reference_id=transfer_id,
                reason=f"Transfer {t['transfer_number']} received",
                user_id=current_user.get("user_id"),
            )

        cur.execute(
            """UPDATE stock_transfers
               SET status='completed', received_by=%s, received_at=NOW()
               WHERE id=%s""",
            (current_user.get("user_id"), transfer_id),
        )
        conn.commit()
        return {"ok": True, "status": "completed"}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.post("/transfers/{transfer_id}/cancel")
def cancel_transfer(transfer_id: int, current_user=Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT * FROM stock_transfers WHERE id=%s FOR UPDATE", (transfer_id,))
        t = cur.fetchone()
        if not t:
            raise HTTPException(status_code=404, detail="Transfer not found")
        if t["status"] != "in_transit":
            raise HTTPException(status_code=400, detail=f"Cannot cancel transfer in status '{t['status']}'")

        cur.execute("SELECT * FROM stock_transfer_items WHERE transfer_id=%s", (transfer_id,))
        items = cur.fetchall()
        for it in items:
            cur.execute("SELECT stock FROM products WHERE id=%s FOR UPDATE", (it["source_product_id"],))
            p = cur.fetchone()
            if not p:
                continue
            add_batch_stock(cur, it["source_product_id"], t["from_branch_id"], int(it["quantity"]), None)
            new_stock = sync_product_from_batches(cur, it["source_product_id"])
            log_movement(
                cur, it["source_product_id"], t["from_branch_id"], "adjustment",
                int(it["quantity"]), new_stock,
                reference_type="transfer_cancel", reference_id=transfer_id,
                reason=f"Transfer {t['transfer_number']} cancelled",
                user_id=current_user.get("user_id"),
            )
        cur.execute(
            "UPDATE stock_transfers SET status='cancelled', cancelled_at=NOW() WHERE id=%s",
            (transfer_id,),
        )
        conn.commit()
        return {"ok": True, "status": "cancelled"}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ─── EXPIRY MANAGEMENT ──────────────────────────────────────────────────────

@router.get("/expiry")
def expiry_report(
    status: str = "near",           # "near" | "expired" | "all"
    days: int = 30,                 # threshold for "near"
    branch_id: Optional[int] = None,
    current_user=Depends(get_current_user),
):
    """Expiry report: stock lots by expiry (supports multiple expiries per product)."""
    from datetime import timedelta
    today = date.today()
    cutoff = today + timedelta(days=days)
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    where = ["p.active = true", "pb.quantity > 0", "pb.expiry_date IS NOT NULL"]
    params: list = []
    if status == "expired":
        where.append("pb.expiry_date < %s")
        params.append(today)
    elif status == "near":
        where.append("pb.expiry_date >= %s AND pb.expiry_date <= %s")
        params += [today, cutoff]
    if current_user.get("role") != "admin":
        ub = current_user.get("branch_id")
        if ub is None:
            raise HTTPException(status_code=403, detail="No branch assigned to this user")
        where.append("p.branch_id = %s")
        params.append(ub)
    elif branch_id is not None:
        where.append("p.branch_id = %s")
        params.append(branch_id)

    w = " AND ".join(where)
    sql = f"""SELECT p.id, p.barcode, p.name_ar, p.name_en, p.category, p.unit,
                     pb.quantity AS stock, p.price, p.cost, pb.expiry_date, p.branch_id,
                     pb.id AS batch_id,
                     b.name_en AS branch_name_en, b.name_ar AS branch_name_ar,
                     (pb.expiry_date - CURRENT_DATE) AS days_left,
                     (pb.quantity * COALESCE(NULLIF(p.cost, 0), p.price, 0)) AS loss_value
              FROM product_batches pb
              JOIN products p ON p.id = pb.product_id
              LEFT JOIN branches b ON p.branch_id = b.id
              WHERE {w}
              ORDER BY pb.expiry_date ASC, p.name_en
              LIMIT %s"""
    count_sql = f"""SELECT COUNT(*)::int AS total
                    FROM product_batches pb
                    JOIN products p ON p.id = pb.product_id
                    WHERE {w}"""
    cur.execute(count_sql, params)
    total_count = int(cur.fetchone()["total"])
    cur.execute(sql, params + [MAX_INVENTORY_ROWS])
    rows = cur.fetchall()
    conn.close()
    return {
        "items": [dict(r) for r in rows],
        "total_count": total_count,
        "shown_count": len(rows),
    }


@router.get("/expiry/summary")
def expiry_summary(days: int = 30,
                   branch_id: Optional[int] = None,
                   current_user=Depends(get_current_user)):
    from datetime import timedelta
    today = date.today()
    cutoff = today + timedelta(days=days)
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    branch_filter = ""
    params: list = []
    if current_user.get("role") != "admin":
        ub = current_user.get("branch_id")
        if ub is None:
            raise HTTPException(status_code=403, detail="No branch assigned to this user")
        branch_filter = " AND p.branch_id = %s"
        params.append(ub)
    elif branch_id is not None:
        branch_filter = " AND p.branch_id = %s"
        params.append(branch_id)

    cur.execute(
        f"""SELECT
              COUNT(*) FILTER (WHERE pb.expiry_date < %s) AS expired_count,
              COALESCE(SUM(pb.quantity * COALESCE(NULLIF(p.cost, 0), p.price, 0))
                FILTER (WHERE pb.expiry_date < %s), 0) AS expired_value,
              COUNT(*) FILTER (WHERE pb.expiry_date >= %s AND pb.expiry_date <= %s) AS near_count,
              COALESCE(SUM(pb.quantity * COALESCE(NULLIF(p.cost, 0), p.price, 0))
                FILTER (WHERE pb.expiry_date >= %s AND pb.expiry_date <= %s), 0) AS near_value
            FROM product_batches pb
            JOIN products p ON p.id = pb.product_id AND p.active = true
            WHERE pb.quantity > 0 AND pb.expiry_date IS NOT NULL{branch_filter}""",
        [today, today, today, cutoff, today, cutoff] + params,
    )
    row = cur.fetchone()
    conn.close()
    return dict(row)


def _branch_stock_data(
    q,
    current_user,
    branch_id: Optional[int] = None,
    keys: Optional[str] = None,
    load_all: bool = False,
):
    """Aggregated per-branch stock balances. Groups products that share the
    same barcode (or, when barcode is empty, the same EN+AR name) and returns
    one row per product family with a breakdown of stock across all branches.
    Admins see every branch; non-admins are restricted to their own branch.
    Optional branch_id limits rows to one branch (admins only)."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        is_admin = current_user.get("role") == "admin"
        user_branch = current_user.get("branch_id")
        if not is_admin and user_branch is None:
            raise HTTPException(status_code=403, detail="No branch assigned to this user")

        effective_branch = user_branch
        if is_admin and branch_id is not None:
            cur.execute("SELECT id FROM branches WHERE id=%s", [branch_id])
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Branch not found")
            effective_branch = branch_id
        elif not is_admin and branch_id is not None and branch_id != user_branch:
            raise HTTPException(status_code=403, detail="Cannot view another branch")

        if is_admin and branch_id is None:
            cur.execute("SELECT id, name_en, name_ar FROM branches ORDER BY id")
        elif effective_branch is not None:
            cur.execute(
                "SELECT id, name_en, name_ar FROM branches WHERE id=%s ORDER BY id",
                [effective_branch],
            )
        else:
            cur.execute("SELECT id, name_en, name_ar FROM branches ORDER BY id")
        branches = [dict(r) for r in cur.fetchall()]

        where = ["p.active = true"]
        params: list = []
        if keys:
            key_list = [k.strip() for k in keys.replace(";", ",").split(",") if k.strip()]
            if key_list:
                where.append(
                    "COALESCE(NULLIF(p.barcode,''), 'name:' || p.name_en || '::' || p.name_ar) = ANY(%s)"
                )
                params.append(key_list)
        if q:
            terms = _parse_search_terms(q)
            if terms:
                term_ors = []
                for term in terms:
                    clause, term_params = _branch_stock_term_clause(term)
                    term_ors.append(clause)
                    params.extend(term_params)
                where.append("(" + " OR ".join(term_ors) + ")")
        elif not load_all and not keys:
            summary = {
                "total_count": 0,
                "shown_count": 0,
                "low_stock": 0,
                "out_of_stock": 0,
                "truncated": False,
            }
            return {"branches": branches, "items": [], "summary": summary}
        if effective_branch is not None:
            where.append("p.branch_id = %s")
            params.append(effective_branch)

        where_sql = " AND ".join(where)
        row_limit = BRANCH_STOCK_LIMIT if (q or keys) else MAX_INVENTORY_ROWS

        grouped_cte = f"""WITH grouped AS (
                  SELECT
                    COALESCE(NULLIF(p.barcode,''), 'name:' || p.name_en || '::' || p.name_ar) AS key,
                    p.branch_id,
                    SUM(p.stock)::int       AS branch_stock,
                    SUM(p.min_stock)::int   AS branch_min,
                    MIN(p.id)               AS product_id,
                    MAX(p.barcode)          AS barcode,
                    MAX(p.international_barcode) AS international_barcode,
                    MAX(p.name_en)          AS name_en,
                    MAX(p.name_ar)          AS name_ar,
                    MAX(p.category)         AS category,
                    MAX(p.unit)             AS unit
                  FROM products p
                  WHERE {where_sql}
                  GROUP BY key, p.branch_id
                )"""

        cur.execute(
            f"""{grouped_cte}
                SELECT
                  COUNT(DISTINCT key)::int AS total_count,
                  COUNT(*) FILTER (
                    WHERE product_id IS NOT NULL AND branch_stock <= 0
                  )::int AS out_of_stock,
                  COUNT(*) FILTER (
                    WHERE product_id IS NOT NULL
                      AND branch_stock > 0 AND branch_stock <= branch_min
                  )::int AS low_stock
                FROM grouped""",
            params,
        )
        summary = dict(cur.fetchone())

        cur.execute(
            f"""{grouped_cte}
                SELECT
                  key,
                  MAX(barcode)   AS barcode,
                  MAX(international_barcode) AS international_barcode,
                  MAX(name_en)   AS name_en,
                  MAX(name_ar)   AS name_ar,
                  MAX(category)  AS category,
                  MAX(unit)      AS unit,
                  SUM(branch_stock)::int AS total_stock,
                  SUM(branch_min)::int   AS total_min,
                  json_agg(json_build_object(
                    'branch_id',  branch_id,
                    'product_id', product_id,
                    'stock',      branch_stock,
                    'min_stock',  branch_min
                  ) ORDER BY branch_id) AS rows
                FROM grouped
                GROUP BY key
                ORDER BY MAX(name_en), key
                LIMIT %s""",
            params + [row_limit],
        )
        items = []
        for r in cur.fetchall():
            d = dict(r)
            by_branch = {row["branch_id"]: row for row in d.pop("rows", []) if row.get("branch_id") is not None}
            d["branches"] = [
                {
                    "branch_id": b["id"],
                    "branch_name_en": b["name_en"],
                    "branch_name_ar": b["name_ar"],
                    "stock": int((by_branch.get(b["id"]) or {}).get("stock") or 0),
                    "min_stock": int((by_branch.get(b["id"]) or {}).get("min_stock") or 0),
                    "product_id": (by_branch.get(b["id"]) or {}).get("product_id"),
                }
                for b in branches
            ]
            items.append(d)
        shown = len(items)
        summary["shown_count"] = shown
        summary["truncated"] = summary["total_count"] > shown
        return {"branches": branches, "items": items, "summary": summary}
    finally:
        cur.close(); conn.close()


@router.get("/branch-stock")
def branch_stock(
    q: Optional[str] = None,
    branch_id: Optional[int] = None,
    keys: Optional[str] = None,
    load_all: bool = False,
    current_user=Depends(get_current_user),
):
    return _branch_stock_data(q, current_user, branch_id, keys, load_all)


@router.get("/branch-stock/export")
def branch_stock_export(
    q: Optional[str] = None,
    branch_id: Optional[int] = None,
    keys: Optional[str] = None,
    load_all: bool = False,
    current_user=Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    def _safe(v):
        if v is None:
            return ""
        s = str(v)
        s = "".join(ch for ch in s if ch in ("\n", "\t") or ord(ch) >= 32)
        s = s.lstrip(" \t\n")
        if s and s[0] in ("=", "+", "-", "@"):
            s = "'" + s
        return s

    data = _branch_stock_data(q, current_user, branch_id, keys, load_all=load_all)
    branches = data["branches"]
    items = data["items"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Branches Stock"
    headers = ["Name (EN)", "Name (AR)", "Barcode", "International Barcode", "Category", "Unit"]
    headers += [b["name_en"] for b in branches]
    headers += ["Total"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1F8A4C")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")

    for row in items:
        by_branch = {b["branch_id"]: b for b in row.get("branches", [])}
        line = [
            _safe(row.get("name_en")),
            _safe(row.get("name_ar")),
            _safe(row.get("barcode")),
            _safe(row.get("international_barcode")),
            _safe(row.get("category")),
            _safe(row.get("unit")),
        ]
        for b in branches:
            cell = by_branch.get(b["id"])
            line.append(int(cell["stock"]) if cell and cell.get("product_id") is not None else 0)
        line.append(int(row.get("total_stock") or 0))
        ws.append(line)

    for col in ws.columns:
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max(width, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="branches_stock.xlsx"'},
    )
